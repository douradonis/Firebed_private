# app.py (ολοκληρωμένο, με ενσωματωμένη λογική για per-line categorization -> epsilon per-vat files)
import os
import sys
import json
import traceback
import logging
from logging.handlers import RotatingFileHandler
import datetime
from typing import Any, List, Dict, Optional
from datetime import datetime as _dt
from werkzeug.utils import secure_filename
from datetime import timezone
from markupsafe import escape, Markup
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify, session
import tempfile
from scraper import scrape_wedoconnect, scrape_mydatapi, scrape_einvoice, scrape_impact, scrape_epsilon
import re
import requests
import pandas as pd
from shutil import move
import importlib
import io
import csv
from scraper_receipt import detect_and_scrape as scrape_receipt
# local mydata helper
from fetch import request_docs
import sys, subprocess, json
from pathlib import Path
# --- Lock + current_app imports (paste here) ---
import threading
try:
    from filelock import FileLock  # προτιμώμενο, cross-process
except Exception:
    # fallback: lightweight in-process lock usable as context manager
    class FileLock:
        def __init__(self, path, timeout=None):
            self.path = path
            self.timeout = timeout
            self._lock = threading.Lock()

        def acquire(self, timeout=None):
            if timeout is None:
                return self._lock.acquire()
            try:
                return self._lock.acquire(timeout=float(timeout))
            except TypeError:
                return self._lock.acquire()

        def release(self):
            try:
                self._lock.release()
            except RuntimeError:
                pass

        def __enter__(self):
            self.acquire(timeout=self.timeout)
            return self

        def __exit__(self, exc_type, exc_val, exc_tb):
            self.release()

from flask import current_app
# --- end lock + current_app imports ---
# --- Compatibility shim: unify invoice-scraper vs receipt-scraper usage ---
# Αυτό το snippet προσπαθεί να χρησιμοποιήσει:
# 1) scrape_receipt από module scraper (αν υπάρχει) ή
# 2) detect_and_scrape από scraper_receipt.py (αν υπάρχει)
# και παρέχει την helper συνάρτηση call_scrape_receipt(mark)
import importlib
scrape_receipt_callable = None

def _normalize_receipt_res(res):
    """Normalize the various possible outputs of the receipt scraper into a dict."""
    try:
        if not res or not isinstance(res, dict):
            return {"ok": False, "error": "scraper returned no dict"}
        return {
            "ok": True,
            "MARK": res.get("MARK") or res.get("mark") or res.get("invoice_id") or res.get("id"),
            "issue_date": res.get("issue_date") or res.get("issueDate") or res.get("date") or "",
            "issuer_vat": res.get("issuer_vat") or res.get("issuer_vat_number") or res.get("issuerAFM") or res.get("AFM") or "",
            "total_amount": res.get("total_amount") or res.get("totalAmount") or res.get("totalValue") or res.get("amount") or "",
            "doc_type": res.get("doc_type") or res.get("docType") or res.get("doc_type_readable") or "",
            "is_invoice": bool(res.get("is_invoice")) if "is_invoice" in res else False,
            "raw": res
        }
    except Exception as e:
        return {"ok": False, "error": f"normalize failed: {e}"}

# try import scrape_receipt from scraper
try:
    _scraper_mod = importlib.import_module("scraper")
    if hasattr(_scraper_mod, "scrape_receipt") and callable(getattr(_scraper_mod, "scrape_receipt")):
        def call_scrape_receipt(mark):
            try:
                res = _scraper_mod.scrape_receipt(mark)
                return _normalize_receipt_res(res)
            except Exception as e:
                return {"ok": False, "error": f"scraper.scrape_receipt failed: {e}"}
        scrape_receipt_callable = call_scrape_receipt
except Exception:
    pass

# fallback to scraper_receipt.detect_and_scrape
if scrape_receipt_callable is None:
    try:
        _sr = importlib.import_module("scraper_receipt")
        if hasattr(_sr, "detect_and_scrape") and callable(getattr(_sr, "detect_and_scrape")):
            def call_scrape_receipt(mark):
                try:
                    res = _sr.detect_and_scrape(mark)
                    return _normalize_receipt_res(res)
                except Exception as e:
                    return {"ok": False, "error": f"scraper_receipt.detect_and_scrape failed: {e}"}
            scrape_receipt_callable = call_scrape_receipt
    except Exception:
        scrape_receipt_callable = None

# Now call_scrape_receipt(mark) is available if either module provided the functionality.


BASE_DIR = os.path.abspath(os.path.dirname(__file__))

DATA_DIR = os.path.join(BASE_DIR, "data")
UPLOADS_DIR = os.path.join(BASE_DIR, "uploads")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOADS_DIR, exist_ok=True)

CACHE_FILE = os.path.join(DATA_DIR, "invoices_cache.json")
SUMMARY_FILE = os.path.join(DATA_DIR, "summary.json")
CREDENTIALS_FILE = os.path.join(DATA_DIR, "credentials.json")
DEFAULT_EXCEL_FILE = os.path.join(UPLOADS_DIR, "invoices.xlsx")
ERROR_LOG = os.path.join(DATA_DIR, "error.log")
EPSILON_JSON_PATH = os.path.join(DATA_DIR, 'epsilon_invoices.json')
EPSILON_EXCEL_PATH = os.path.join(DATA_DIR, 'epsilon_invoices.xlsx')
MARK_COUNTER_PATH = os.path.join(DATA_DIR, 'mark_counter.json')

AADE_USER_ENV = os.getenv("AADE_USER_ID", "")
AADE_KEY_ENV = os.getenv("AADE_SUBSCRIPTION_KEY", "")
MYDATA_ENV = (os.getenv("MYDATA_ENV") or "sandbox").lower()
ALLOWED_CLIENT_EXT = {'.xlsx', '.xls', '.csv'}



app = Flask(__name__, template_folder=TEMPLATES_DIR)
app.secret_key = os.getenv("FLASK_SECRET", "change-me")
app.config["UPLOAD_FOLDER"] = UPLOADS_DIR
FISCAL_META = 'fiscal.meta.json'   # αποθηκεύεται μέσα στο DATA_DIR
REQUIRED_CLIENT_COLUMNS = {"ΑΦΜ", "Επωνυμία", "Διεύθυνση", "Πόλη", "ΤΚ", "Τηλέφωνο"}  # προσάρμοσε αν χρειάζεται




@app.before_request
def log_request_path():
    log.info("Incoming request: method=%s path=%s remote=%s ref=%s", request.method, request.path, request.remote_addr, request.referrer)
# Logging (unchanged)
log = logging.getLogger("mydata_app")
log.setLevel(logging.INFO)
if not log.handlers:
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    log.addHandler(sh)

    fh = RotatingFileHandler(ERROR_LOG, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8")
    fh.setLevel(logging.INFO)
    fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s\n"))
    log.addHandler(fh)

log.info("Starting app - MYDATA_ENV=%s", MYDATA_ENV)
GLOBAL_ACCOUNTS_NAME = "__global_accounts__"
# keep settings inside data/ so it's co-located with other app files
SETTINGS_FILE = os.path.join(DATA_DIR, "credentials_settings.json")
VAT_MAP = {
    "1": "ΦΠΑ 24%",
    "2": "ΦΠΑ 13%",
    "3": "ΦΠΑ 6%",
    "4": "ΦΠΑ 17%",
    "5": "ΦΠΑ 9%",
    "6": "ΦΠΑ 4%",
    "7": "Εξαιρούμενο άρθρο 39α",
    "8": "Εξαιρούμενο άρθρο 47β",
    "9": "Άνευ ΦΠΑ",
}

# --- helper: parse year from a variety of date strings ---
APP_DIR = Path(__file__).resolve().parent
SCRAPER_PATH = APP_DIR / "scraper_receipt.py"
def save_receipt(afm: str, year: str, summary: dict, lines: list, active_years: list):
    """
    Αποθηκεύει μια απόδειξη στα αρχεία AFM_invoices.xlsx και AFM_epsilon_invoices.json
    με τον ίδιο τρόπο που αποθηκεύονται τα τιμολόγια.
    """

    # Έλεγχος ενεργής χρήσης
    issue_date = datetime.strptime(summary["issueDate"], "%d/%m/%Y")
    if int(year) not in active_years or issue_date.year != int(year):
        return {"ok": False, "error": "Απόδειξη εκτός ενεργής χρήσης"}

    # Προετοιμασία φακέλων/αρχειων
    invoices_file = f"data/{afm}_invoices.xlsx"
    epsilon_file = f"data/epsilon/{afm}_epsilon_invoices.json"
    os.makedirs(os.path.dirname(invoices_file), exist_ok=True)
    os.makedirs(os.path.dirname(epsilon_file), exist_ok=True)

    # Προσθήκη τύπου "αποδειξακια"
    summary["type"] = "αποδειξακια"
    for line in lines:
        line["type"] = "αποδειξακια"

    # --- Ενημέρωση JSON ---
    epsilon_data = []
    if os.path.exists(epsilon_file):
        with open(epsilon_file, "r", encoding="utf-8") as f:
            try:
                epsilon_data = json.load(f)
            except:
                epsilon_data = []

    epsilon_data.append({"summary": summary, "lines": lines})

    with open(epsilon_file, "w", encoding="utf-8") as f:
        json.dump(epsilon_data, f, ensure_ascii=False, indent=2)

    # --- Ενημέρωση Excel ---
    df_summary = pd.DataFrame([summary])
    df_lines = pd.DataFrame(lines)

    if os.path.exists(invoices_file):
        with pd.ExcelWriter(invoices_file, mode="a", if_sheet_exists="overlay", engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="summary", index=False, header=False, startrow=writer.sheets["summary"].max_row)
            df_lines.to_excel(writer, sheet_name="lines", index=False, header=False, startrow=writer.sheets["lines"].max_row)
    else:
        with pd.ExcelWriter(invoices_file, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="summary", index=False)
            df_lines.to_excel(writer, sheet_name="lines", index=False)

    return {"ok": True}

def run_scraper_subprocess(arg=None, timeout=240):
    """
    Καλεί scraper_receipt.py με τον ίδιο python interpreter (sys.executable).
    Επιστρέφει dict: { ok: bool, data: dict|None, error: str|None, stdout, stderr }
    """
    python_exec = sys.executable or "python3"
    cmd = [python_exec, str(SCRAPER_PATH)]
    if arg is not None:
        cmd.append(str(arg))
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout)
    except Exception as e:
        return {"ok": False, "data": None, "error": f"subprocess_failed:{e}", "stdout": "", "stderr": ""}

    stdout = (proc.stdout or "").strip()
    stderr = (proc.stderr or "").strip()
    if proc.returncode != 0:
        return {"ok": False, "data": None, "error": f"scraper_exit_{proc.returncode}", "stdout": stdout, "stderr": stderr}

    # proc.returncode == 0 -> try parse stdout as JSON
    if not stdout:
        return {"ok": False, "data": None, "error": "empty_stdout_from_scraper", "stdout": "", "stderr": stderr}
    try:
        parsed = json.loads(stdout)
        return {"ok": True, "data": parsed, "error": None, "stdout": stdout, "stderr": stderr}
    except Exception:
        # scraper completed but didn't emit JSON — return raw stdout for debugging
        return {"ok": False, "data": None, "error": "invalid_json_from_scraper", "stdout": stdout, "stderr": stderr}

def get_excel_path(afm, year):
    return os.path.join(BASE_DIR, f"{afm}_{year}_invoices.xlsx")

def get_json_path(afm):
    return os.path.join(BASE_DIR, f"{afm}_epsilon_invoices.json")

def load_json(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_excel(path):
    if os.path.exists(path):
        return load_workbook(path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["MARK", "Issue Date", "Total Amount", "Type"])  # header
        return wb

def save_excel(wb, path):
    wb.save(path)

def mark_matches_year(receipt, year):
    """Check if issue_date matches selected year"""
    try:
        issue_year = dt.strptime(receipt["issue_date"], "%d/%m/%Y").year
        return issue_year == int(year)
    except:
        return False
def parse_year_from_date_string(s):
    if not s:
        return None
    s = str(s).strip()
    # try common date formats
    patterns = [
        '%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y',
        '%d.%m.%Y', '%Y.%m.%d'
    ]
    for p in patterns:
        try:
            dt = datetime.strptime(s, p)
            return dt.year
        except Exception:
            pass
    # fallback: find any 4-digit year
    m = re.search(r'(19|20)\d{2}', s)
    if m:
        return int(m.group(0))
    return None

# --- helper: persistent 15-digit MARK generator ---
def get_next_mark():
    """
    Read/update MARK_COUNTER_PATH to return next integer as 15-digit zero-padded string.
    Keeps a single counter. Thread/process safe-ish by using FileLock.
    """
    lock_path = MARK_COUNTER_PATH + '.lock'
    lock = FileLock(lock_path, timeout=5)
    with lock:
        try:
            if os.path.exists(MARK_COUNTER_PATH):
                with open(MARK_COUNTER_PATH, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                data = {'last': 0}
        except Exception:
            data = {'last': 0}
        last = int(data.get('last', 0) or 0)
        nxt = last + 1
        data['last'] = nxt
        with open(MARK_COUNTER_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f)
    return str(nxt).zfill(15)

# --- helper: append receipt entry to Excel (pandas) with file lock ---
def append_receipt_to_excel(entry, excel_path=EPSILON_EXCEL_PATH):
    lock_path = excel_path + '.lock'
    lock = FileLock(lock_path, timeout=10)
    with lock:
        # normalize columns we want to store (flat)
        flat = {
            'MARK': entry.get('MARK'),
            'saved_at': entry.get('saved_at'),
            'url': entry.get('url'),
            'issuer_vat': entry.get('issuer_vat'),
            'issuer_name': entry.get('issuer_name'),
            'issue_date': entry.get('issue_date'),
            'progressive_aa': entry.get('progressive_aa'),
            'total_amount': entry.get('total_amount')
        }
        # include raw as JSON string optionally
        flat['raw'] = json.dumps(entry.get('raw', {}), ensure_ascii=False)
        if os.path.exists(excel_path):
            try:
                df_existing = pd.read_excel(excel_path, dtype=str)
            except Exception:
                df_existing = pd.DataFrame()
            df_new = pd.DataFrame([flat])
            df_combined = pd.concat([df_existing, df_new], ignore_index=True, sort=False)
            df_combined.to_excel(excel_path, index=False)
        else:
            pd.DataFrame([flat]).to_excel(excel_path, index=False)
def get_existing_client_ids() -> set:
    """
    Return a set of existing client IDs (ΑΦΜ) from current client_db (if any).
    Falls back to empty set if no client_db exists.
    """
    client_ids = set()
    try:
        # αναζήτηση τρέχοντος client_db
        for existing in os.listdir(DATA_DIR):
            if existing.startswith('client_db') and os.path.splitext(existing)[1].lower() in ALLOWED_CLIENT_EXT:
                path = os.path.join(DATA_DIR, existing)
                ext = os.path.splitext(path)[1].lower()
                if ext in ['.xls', '.xlsx']:
                    df = pd.read_excel(path, dtype=str)
                else:
                    df = pd.read_csv(path, dtype=str)
                df.fillna('', inplace=True)
                for afm in df.get("ΑΦΜ", []):
                    afm_str = str(afm).strip()
                    if afm_str:
                        client_ids.add(afm_str)
                break  # παίρνουμε μόνο το πρώτο υπάρχον client_db
    except Exception:
        log.exception("Failed to get existing client IDs from client_db")
    return client_ids

def _get_mark_from_epsilon_item(item):
    # normalize possible keys that may hold the mark
    for k in ("mark", "MARK", "invoice_id", "Αριθμός Μητρώου", "id"):
        if k in item and item.get(k) not in (None, ""):
            return str(item.get(k)).strip()
    return ""

def sync_epsilon_with_excel(vat):
    """
    Sync per-vat epsilon cache with the Excel file for vat:
      - Keep only epsilon entries whose mark exists in the Excel MARK column.
      - If Excel exists but has zero rows (no MARKs), epsilon will be truncated to [].
    Returns a tuple (changed: bool, removed_count: int).
    """
    try:
        excel_path = excel_path_for(vat=vat)
        if not os.path.exists(excel_path):
            log.info("sync_epsilon_with_excel: excel not found for vat %s -> skipping sync", vat)
            return False, 0

        try:
            df = pd.read_excel(excel_path, engine="openpyxl", dtype=str).fillna("")
        except Exception as e:
            log.exception("sync_epsilon_with_excel: failed reading excel %s", excel_path)
            return False, 0

        if "MARK" in df.columns:
            marks_in_excel = set(df["MARK"].astype(str).str.strip().tolist())
        else:
            # no MARK column -> treat as empty set (remove all)
            marks_in_excel = set()

        eps_list = load_epsilon_cache_for_vat(vat) or []
        # keep only those whose normalized mark is present in marks_in_excel
        kept = []
        removed = []
        for it in eps_list:
            try:
                m = _get_mark_from_epsilon_item(it)
                if m and m in marks_in_excel:
                    kept.append(it)
                else:
                    removed.append(it)
            except Exception:
                # if something weird, keep item (safer) OR you can choose to remove; here we keep
                kept.append(it)

        if len(removed) == 0:
            log.debug("sync_epsilon_with_excel: nothing to remove for vat %s (marks kept=%d)", vat, len(kept))
            return False, 0

        # persist truncated cache
        try:
            _safe_save_epsilon_cache(vat, kept)
            log.info("sync_epsilon_with_excel: removed %d epsilon entries for vat %s (kept=%d)", len(removed), vat, len(kept))
            return True, len(removed)
        except Exception:
            log.exception("sync_epsilon_with_excel: failed saving epsilon after sync for vat %s", vat)
            return False, 0

    except Exception:
        log.exception("sync_epsilon_with_excel: unexpected error for vat %s", vat)
        return False, 0

def create_empty_excel_for_vat(vat, fiscal_year=None):
    """Create an empty excel file with standard headers for given vat + fiscal_year."""
    try:
        safe_vat = secure_filename(str(vat))
    except Exception:
        safe_vat = str(vat)

    # prefer get_active_fiscal_year if fiscal_year not passed
    if fiscal_year is None:
        getter = globals().get("get_active_fiscal_year")
        try:
            if callable(getter):
                fiscal_year = getter()
        except Exception:
            fiscal_year = None
    if fiscal_year is None:
        from datetime import datetime
        fiscal_year = datetime.now().year

    # Use excel_path_for to keep filename consistent
    excel_path = excel_path_for(vat=vat)
    if os.path.exists(excel_path):
        log.debug("create_empty_excel_for_vat: excel already exists: %s", excel_path)
        return excel_path

    cols = [
        "MARK", "ΑΦΜ", "Επωνυμία", "Σειρά", "Αριθμός",
        "Ημερομηνία", "Είδος", "ΦΠΑ_ΚΑΤΗΓΟΡΙΑ",
        "Καθαρή Αξία", "ΦΠΑ", "Σύνολο"
    ]
    import pandas as pd
    df = pd.DataFrame(columns=cols).astype(str)
    try:
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        df.to_excel(excel_path, index=False, engine="openpyxl")
        log.info("create_empty_excel_for_vat: created empty excel %s", excel_path)
    except Exception:
        log.exception("create_empty_excel_for_vat: failed creating excel %s", excel_path)
        raise
    return excel_path






import re
def _fiscal_meta_path():
    """Return path to fiscal meta file inside DATA_DIR."""
    return os.path.join(DATA_DIR, "fiscal_meta.json")
def epsilon_item_has_detail(item):
    """
    True αν το item φαίνεται 'πραγματικό' (περιέχει αρκετά πεδία).
    False αν είναι placeholder (π.χ. μόνο mark/AFM/AA + empty lines).
    Κανόνες:
      - issueDate OR
      - lines με πραγματικά πεδία (description/amount/vat) OR
      - totalNetValue/totalValue OR
      - AFM_issuer + (AA ή issueDate ή totalValue)
    """
    try:
        if not item or not isinstance(item, dict):
            return False
        if item.get("issueDate"):
            return True
        if item.get("totalNetValue") or item.get("totalValue"):
            return True
        lines = item.get("lines") or []
        if isinstance(lines, list):
            for l in lines:
                if l and (l.get("description") or l.get("amount") or l.get("vat")):
                    return True
        # AFM_issuer alone is not enough; require some other info
        if item.get("AFM_issuer") or item.get("AFM"):
            if item.get("aa") or item.get("AA") or item.get("issueDate") or item.get("totalValue"):
                return True
    except Exception:
        pass
    return False



# --- Ensure _safe_save_epsilon_cache exists (put this near top of app.py, after imports) ---
def _safe_save_epsilon_cache(vat_code, epsilon_list):
    """
    Compatible safe writer used across the app.
    Returns the path of the saved file on success.
    Raises on failure.
    """
    epsilon_dir = os.path.join(DATA_DIR, "epsilon")
    os.makedirs(epsilon_dir, exist_ok=True)
    safe_vat = secure_filename(str(vat_code))
    epsilon_path = os.path.join(epsilon_dir, f"{safe_vat}_epsilon_invoices.json")

    # coerce to list
    if epsilon_list is None:
        epsilon_list = []
    if not isinstance(epsilon_list, list):
        if isinstance(epsilon_list, dict):
            epsilon_list = [epsilon_list]
        else:
            try:
                epsilon_list = list(epsilon_list)
            except Exception:
                epsilon_list = [epsilon_list]

    # quick sanity: avoid overwriting with mostly-placeholders if file exists
    incomplete = 0
    try:
        for it in epsilon_list:
            if not (it and (it.get("lines") or it.get("issueDate") or it.get("AFM_issuer") or it.get("AFM"))):
                incomplete += 1
    except Exception:
        incomplete = 0

    try:
        if len(epsilon_list) > 0 and incomplete >= max(1, int(len(epsilon_list) * 0.9)):
            if os.path.exists(epsilon_path):
                log.warning("_safe_save_epsilon_cache: skipping overwrite (looks like placeholders) %s", epsilon_path)
                return epsilon_path
    except Exception:
        pass

    # Try to use json_write if defined; otherwise do atomic tmp write here
    try:
        # prefer json_write helper if present
        if 'json_write' in globals() and callable(globals().get('json_write')):
            json_write(epsilon_path, epsilon_list)
            try:
                log.info("_safe_save_epsilon_cache: saved epsilon to %s", epsilon_path)
            except Exception:
                pass
            return epsilon_path
    except Exception:
        log.exception("_safe_save_epsilon_cache: json_write failed, falling back to direct atomic write")

    # fallback atomic write
    tmp = None
    try:
        text = json.dumps(epsilon_list, ensure_ascii=False, indent=2)
        fd, tmp = tempfile.mkstemp(prefix=".tmp_epsilon_", dir=epsilon_dir)
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            fh.write(text)
            fh.flush()
            try:
                os.fsync(fh.fileno())
            except Exception:
                pass
        os.replace(tmp, epsilon_path)
        try:
            log.info("_safe_save_epsilon_cache: saved epsilon (fallback) to %s", epsilon_path)
        except Exception:
            pass
        return epsilon_path
    except Exception:
        try:
            log.exception("_safe_save_epsilon_cache: fallback write failed")
        except Exception:
            print("fallback write failed for", epsilon_path)
        # cleanup tmp
        try:
            if tmp and os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        raise
# --- end _safe_save_epsilon_cache ---




def get_active_fiscal_year():
    """
    Read persisted fiscal year (int) from DATA_DIR/fiscal_meta.json.
    Returns integer fiscal year or None if not found / on error.
    """
    try:
        p = _fiscal_meta_path()
        if not os.path.exists(p):
            return None
        with open(p, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        if not data:
            return None
        fy = data.get("fiscal_year")
        if fy is None:
            return None
        try:
            return int(fy)
        except Exception:
            # stored value not an int
            return None
    except Exception:
        try:
            log.exception("get_active_fiscal_year: failed to read fiscal meta")
        except Exception:
            pass
        return None


def set_active_fiscal_year(year):
    """Persist fiscal year (int). Returns True on success, False on failure."""
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        p = _fiscal_meta_path()
        with open(p, "w", encoding="utf-8") as fh:
            json.dump({"fiscal_year": int(year)}, fh)
        try:
            log.info("Set active fiscal year: %s", year)
        except Exception:
            pass
        return True
    except Exception:
        try:
            log.exception("Failed to write fiscal meta")
        except Exception:
            pass
        return False
# ---------------- normalize helper (paste/replace existing) ----------------
def _normalize_afm(raw):
    """Καθαρίζει AFM: κρατά μόνο digits, κόβει περιττά και επιστρέφει None αν άδειο."""
    if not raw:
        return None
    s = str(raw).strip()
    # extract digits
    digits = re.sub(r'\D', '', s)
    if not digits:
        return None
    # common AFM length = 9, but αν έχει περισσότερα κρατάμε τα *πρώτα* 9 (αν αυτό θέλεις)
    if len(digits) > 9:
        digits = digits[-9:]  # καλύτερο να πάρουμε τα **τελευταία 9** (συχνά οι σελίδες έχουν πρόσθετα prefix)
    return digits
# --- νέο helper: update or append a summary row into excel ---
def _normalize_summary_row_for_excel(summary):
    """Return dict with keys aligned to DEFAULT EXCEL columns used in your app."""
    return {
        "MARK": str(summary.get("mark","") or ""),
        "AA": str(summary.get("AA","") or ""),
        "AFM": str(summary.get("AFM","") or ""),
        "Name": str(summary.get("Name","") or ""),
        "issueDate": str(summary.get("issueDate","") or ""),
        "totalValue": str(summary.get("totalValue","") or ""),
        "category": str(summary.get("category","") or ""),
        "note": str(summary.get("note","") or ""),
        "created_at": str(summary.get("created_at","") or ""),
    }

def _ensure_excel_and_update_or_append(summary: dict, vat: Optional[str] = None, cred_name: Optional[str] = None):
    """
    Ensure an excel exists (create if missing) and either update existing row matching mark or append new row.
    Uses pandas to load/edit/save — simpler and reliable for small files.
    """
    path = excel_path_for(cred_name=cred_name, vat=vat)
    # headers consistent with DEFAULT_EXCEL_FILE usage above
    headers = ["MARK","AA","AFM","Name","issueDate","totalValue","category","note","created_at"]

    # create file if missing with headers
    if not os.path.exists(path):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        df = pd.DataFrame(columns=headers)
        df.to_excel(path, index=False)
    # load existing
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception:
        # fallback: make empty DF with headers
        df = pd.DataFrame(columns=headers)

    # normalize types: ensure all header cols exist
    for h in headers:
        if h not in df.columns:
            df[h] = ""

    row = _normalize_summary_row_for_excel(summary)
    mark_val = str(row.get("MARK","")).strip()

    # try to find existing row by mark (use _match_row_by_mark logic)
    found = False
    if mark_val:
        # use vectorized comparison across likely colnames
        mask = pd.Series([False]*len(df), index=df.index)
        # prefer 'MARK' column, else try any object/string column
        if "MARK" in df.columns:
            try:
                mask = df["MARK"].fillna("").astype(str).str.strip() == mark_val
            except Exception:
                mask = (df["MARK"].fillna("") == mark_val)
        if not mask.any():
            # scan other string columns
            for c in df.columns:
                try:
                    if df[c].dtype == 'O' or df[c].dtype == 'string':
                        m2 = df[c].fillna("").astype(str).str.strip() == mark_val
                        if m2.any():
                            mask = m2
                            break
                except Exception:
                    continue

        if mask.any():
            # update first match
            idx = mask[mask].index[0]
            for h in headers:
                # don't overwrite MARK,AA if blank in new row? We'll overwrite with new data
                df.at[idx, h] = row.get(h,"")
            found = True

    if not found:
        # append new row
        df = df.append(row, ignore_index=True)

    # ensure AFM and MARK are saved as text in Excel: pandas will save as strings; that's OK
    try:
        df.to_excel(path, index=False)
    except Exception:
        # final fallback - try openpyxl write: create workbook with headers then append row
        try:
            from openpyxl import Workbook, load_workbook
            if not os.path.exists(path):
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
                wb.save(path)
            wb = load_workbook(path)
            ws = wb.active
            # if updated, easier to rewrite entire excel from df
            for r in dataframe_to_rows(df, index=False, header=True):
                pass
            # simpler: use pandas to_excel in a temp file and replace
            tmp = path + ".tmp.xlsx"
            df.to_excel(tmp, index=False)
            os.replace(tmp, path)
        except Exception:
            log.exception("Failed to write excel for summary for mark=%s", mark_val)

def _extract_headers_from_upload(file_stream, ext):
    """
    Προσπαθεί να διαβάσει τα headers από το uploaded file_stream.
    - file_stream: κινούμενο binary stream στη θέση αρχής (file.read()-compatible).
    - ext: '.xlsx' / '.xls' / '.csv'
    Επιστρέφει (success: bool, headers: list[str] or None, error_msg: str or None)
    """
    # ensure stream at start
    try:
        file_stream.seek(0)
    except Exception:
        pass

    if ext == '.csv':
        # fallback χωρίς pandas: διαβάζουμε μόνο την πρώτη γραμμή
        try:
            text = file_stream.read().decode('utf-8-sig')  # handle BOM
            # move back in case caller wants to re-read
            file_stream.seek(0)
            reader = csv.reader(io.StringIO(text))
            first = next(reader, None)
            if first is None:
                return False, None, 'Το CSV φαίνεται άδειο.'
            headers = [h.strip() for h in first]
            return True, headers, None
        except Exception as e:
            return False, None, f'Σφάλμα κατά την ανάγνωση CSV headers: {e}'
    else:
        # προσπαθούμε με pandas για excel/xls
        try:
            # pandas θα διαβάσει μόνο τα headers (nrows=0) — γρήγορο
            file_stream.seek(0)
            df = pd.read_excel(file_stream, nrows=0, engine='openpyxl' if ext == '.xlsx' else None)
            headers = [str(h).strip() for h in df.columns.tolist()]
            file_stream.seek(0)
            return True, headers, None
        except Exception as e:
            # αν pandas δεν εγκατεστημένο ή άλλο σφάλμα
            return False, None, f'Αποτυχία ανάγνωσης Excel (pandas/openpyxl απαιτείται): {e}'
def _normalize_receipt_summary(summary: dict) -> dict:
    s = dict(summary or {})
    out = {}
    out['MARK'] = (s.get('MARK') or s.get('mark') or '').strip()
    out['AA'] = (s.get('AA') or s.get('progressive_aa') or s.get('id') or '').strip()
    # prefer issuer_vat, but normalize to AFM field (9 digits)
    raw_vat = (s.get('issuer_vat') or s.get('AFM') or s.get('vat') or s.get('issuerVat') or '')
    out['AFM'] = _normalize_afm(raw_vat) or ''
    out['issuer_vat_raw'] = (raw_vat or '').strip()
    out['Name'] = (s.get('issuer_name') or s.get('Name') or s.get('company') or '').strip()
    # date normalization: keep raw (you may reformat if you want)
    out['issueDate'] = (s.get('issue_date') or s.get('issueDate') or s.get('date') or '').strip()
    # total amount normalization: reuse your existing _clean_amount_to_comma if present
    total = s.get('total_amount') or s.get('totalValue') or s.get('total') or ''
    try:
        # try to keep same formatting used in invoices (comma decimal)
        out['totalValue'] = _clean_amount_to_comma(total) or str(total)
    except Exception:
        out['totalValue'] = str(total)
    out['type_name'] = (s.get('type_name') or s.get('doc_type') or 'Απόδειξη').strip()
    out['lines'] = s.get('lines') or []
    out['_saved_at'] = datetime.utcnow().isoformat() + 'Z'
    # keep category if any line already had it
    out['category'] = ''
    for ln in out['lines']:
        if isinstance(ln, dict) and ln.get('category'):
            out['category'] = ln.get('category')
            break
    return out


def _normalize_key_val(x):
    try:
        if x is None:
            return ""
        # convert numbers to str, strip whitespace, lower-case for robust comparison
        return str(x).strip()
    except Exception:
        return ""

def _find_afm_in_epsilon(mark: str = None, aa: str = None) -> str:
    """
    Search data/epsilon/*.json for an invoice matching mark or AA.
    Return AFM_issuer or AFM if found, else empty string.
    """
    try:
        epsilon_dir = os.path.join(DATA_DIR, "epsilon")
        if not os.path.isdir(epsilon_dir):
            return ""
        for fname in os.listdir(epsilon_dir):
            if not fname.endswith("_epsilon_invoices.json"):
                continue
            try:
                path = os.path.join(epsilon_dir, fname)
                with open(path, "r", encoding="utf-8") as fh:
                    items = json.load(fh)
                if not isinstance(items, list):
                    continue
                for it in items:
                    try:
                        if mark and str(it.get("mark", "")).strip() and str(it.get("mark", "")).strip() == str(mark).strip():
                            return (it.get("AFM_issuer") or it.get("AFM") or "").strip()
                        if aa and str(it.get("AA", "")).strip() and str(it.get("AA", "")).strip() == str(aa).strip():
                            return (it.get("AFM_issuer") or it.get("AFM") or "").strip()
                    except Exception:
                        continue
            except Exception:
                # ignore corrupt epsilon files
                continue
    except Exception:
        try:
            log.exception("_find_afm_in_epsilon failed")
        except Exception:
            pass
    return ""



def _append_to_excel(rec_dict, vat: Optional[str] = None, cred_name: Optional[str] = None):
    """
    Append a single record as a row to the per-vat Excel file determined by excel_path_for().
    Uses AFM from rec_dict first, otherwise looks into epsilon cache (by mark / AA).
    """
    # determine vat to use for per-vat excel filename (fallbacks)
    vat_candidate = vat or rec_dict.get('issuer_vat') or rec_dict.get('issuer_vat_raw') or rec_dict.get('AFM') or ""
    safe_vat = secure_filename(str(vat_candidate)) if vat_candidate else ""
    path = excel_path_for(cred_name=cred_name, vat=safe_vat)

    # ensure directory exists
    os.makedirs(os.path.dirname(path), exist_ok=True)

    # Try to resolve AFM: prefer explicit issuer fields, else search epsilon cache by mark/AA
    afm = (
        (rec_dict.get('issuer_vat') or rec_dict.get('AFM_issuer') or rec_dict.get('AFM') or rec_dict.get('issuer_vat_raw') or "")
        .strip()
    )
    if not afm:
        # attempt to find AFM in epsilon cache using mark / AA
        mark = rec_dict.get("MARK") or rec_dict.get("mark") or rec_dict.get("mark_id") or ""
        aa = rec_dict.get("AA") or rec_dict.get("aa") or rec_dict.get("invoice_number") or ""
        found = _find_afm_in_epsilon(mark=mark, aa=aa)
        if found:
            afm = found.strip()

    # build row (add/adjust fields as your app expects)
    row = {
        'saved_at': rec_dict.get('_saved_at'),
        'MARK': rec_dict.get('MARK') or rec_dict.get('mark'),
        'AA': rec_dict.get('AA') or rec_dict.get('aa'),
        'AFM': afm,
        'Name': rec_dict.get('Name') or rec_dict.get('Name_issuer') or rec_dict.get('Name_counterparty') or "",
        'issueDate': rec_dict.get('issueDate') or rec_dict.get('issueDate_raw') or rec_dict.get('issue_date') or "",
        'totalValue': rec_dict.get('totalValue') or rec_dict.get('total_value') or rec_dict.get('total') or "",
        'category': rec_dict.get('category') or rec_dict.get('classification') or ""
    }

    headers = list(row.keys())

    # debug log
    try:
        log.info("append_to_excel -> path=%s vat_candidate=%s resolved_afm=%s mark=%s AA=%s",
                 path, safe_vat, row['AFM'], row.get('MARK'), row.get('AA'))
    except Exception:
        pass

    # Try pandas path first (preferred)
    try:
        import pandas as pd
        if os.path.exists(path):
            df_existing = pd.read_excel(path, engine='openpyxl', dtype=str)
        else:
            df_existing = pd.DataFrame(columns=headers)

        df_new = pd.DataFrame([row])
        df_concat = pd.concat([df_existing, df_new], ignore_index=True, sort=False)

        # Ensure all headers exist (order)
        for h in headers:
            if h not in df_concat.columns:
                df_concat[h] = ""

        df_concat.to_excel(path, index=False, engine='openpyxl')
        return True

    except Exception as e_pandas:
        # Fallback to openpyxl direct append/create
        try:
            from openpyxl import load_workbook, Workbook
            if not os.path.exists(path):
                wb = Workbook()
                ws = wb.active
                ws.append(headers)
                ws.append([row.get(k, '') for k in headers])
                wb.save(path)
                return True

            wb = load_workbook(path)
            ws = wb.active
            # ensure header row exists and matches headers; if not, add header if missing
            existing_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if not existing_headers or all(h is None for h in existing_headers):
                # insert header then continue
                ws.insert_rows(1)
                for idx, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=idx, value=h)
            ws.append([row.get(k, '') for k in headers])
            wb.save(path)
            return True

        except Exception as e_openpyxl:
            try:
                log.exception("append_to_excel failed (pandas err=%s, openpyxl err=%s)", e_pandas, e_openpyxl)
            except Exception:
                print("append_to_excel failed:", e_pandas, e_openpyxl)
            return False
    """
    Append a single record as a row to EXCEL_FILE.
    Columns: saved_at, MARK, AA, AFM, Name, issueDate, totalValue, category
    Uses pandas (openpyxl engine).
    """
    row = {
        'saved_at': rec_dict.get('_saved_at'),
        'MARK': rec_dict.get('MARK'),
        'AA': rec_dict.get('AA'),
        'AFM': rec_dict.get('AFM'),
        'Name': rec_dict.get('Name'),
        'issueDate': rec_dict.get('issueDate'),
        'totalValue': rec_dict.get('totalValue'),
        'category': rec_dict.get('category') or ''
    }
    try:
        if os.path.exists(EXCEL_FILE):
            df_existing = pd.read_excel(EXCEL_FILE, engine='openpyxl')
        else:
            df_existing = pd.DataFrame(columns=list(row.keys()))
        df_new = pd.DataFrame([row])
        df_concat = pd.concat([df_existing, df_new], ignore_index=True, sort=False)
        df_concat.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
        return True
    except Exception as e:
        current_app.logger.exception("excel append failed: %s", e)
        return False


def set_active_fiscal_year(year):
    """Persist fiscal year (int)."""
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        p = _fiscal_meta_path()
        with open(p, 'w', encoding='utf-8') as fh:
            json.dump({'fiscal_year': int(year)}, fh)
        log.info("Set active fiscal year: %s", year)
        return True
    except Exception:
        log.exception("Failed to write fiscal meta")
        return False
def _client_meta_path():
    """Full path to metadata JSON for client_db inside DATA_DIR."""
    return os.path.join(DATA_DIR, 'client_db.meta.json')

def read_client_meta():
    """Read metadata if exists, return dict or None."""
    meta_path = _client_meta_path()
    try:
        if os.path.exists(meta_path):
            with open(meta_path, 'r', encoding='utf-8') as fh:
                return json.load(fh)
    except Exception:
        log.exception("Failed reading client_db.meta.json")
    return None

def write_client_meta(filename, uploaded_at_iso):
    """Write metadata (filename, uploaded_at iso) to meta file."""
    meta = {
        'filename': filename,
        'uploaded_at': uploaded_at_iso
    }
    meta_path = _client_meta_path()
    try:
        with open(meta_path, 'w', encoding='utf-8') as fh:
            json.dump(meta, fh, ensure_ascii=False, indent=2)
    except Exception:
        log.exception("Failed writing client_db.meta.json")

def normalize_vat_key(raw):
    """
    Normalize different vat-category representations into a canonical form like '24%','0%','6%','13%'
    Examples:
      'ΦΠΑ 24%' -> '24%'
      '24%' -> '24%'
      '24' -> '24%'
      'VAT 24%' -> '24%'
      '0%' -> '0%'
    Returns empty string for unknown/empty.
    """
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    if not s:
        return ""
    # remove common words like 'φπα' or 'vat'
    s = re.sub(r'φπα|\bvat\b', '', s, flags=re.IGNORECASE).strip()
    # find a number (integer or decimal) optionally followed by %
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*%?', s)
    if m:
        num = m.group(1).replace(',', '.')
        # if it's integer-like, keep integer
        if '.' in num:
            # keep as-is (rare), but normalize trailing .0
            try:
                f = float(num)
                if f.is_integer():
                    num = str(int(f))
                else:
                    # keep one or two decimals? keep as original trimmed
                    num = num.rstrip('0').rstrip('.') if '.' in num else num
            except:
                num = num
        # canonical form: without decimals if integer, with '%' suffix
        return f"{num}%"
    # if there is something else like 'μηδεν' -> map to 0%
    if re.search(r'0|μηδ', s):
        return "0%"
    return ""
def read_credentials_list():
    try:
        with open(CREDENTIALS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return []
    except Exception:
        log.exception("read_credentials_list failed")
        return []

def write_credentials_list(data_list):
    tmp = CREDENTIALS_FILE + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data_list, f, ensure_ascii=False, indent=2)
    os.replace(tmp, CREDENTIALS_FILE)

def find_active_client_index(creds_list, vat=None):
    # priority: match vat, else find 'active': true, else first
    if vat:
        for i,c in enumerate(creds_list):
            if str(c.get('vat','')).strip() == str(vat).strip():
                return i
    for i,c in enumerate(creds_list):
        if c.get('active'):
            return i
    return 0 if creds_list else None

def load_invoices():
    if os.path.exists(INVOICES_FILE):
        with open(INVOICES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

# φορτώνουμε epsilon_invoices.json
def load_epsilon():
    if os.path.exists(EPSILON_FILE):
        with open(EPSILON_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []
def get_global_accounts_from_credentials() -> Dict:
    """
    Διαβάζει το credentials.json και επιστρέφει το αντικείμενο accounts
    αν υπάρχει ως credential με name == GLOBAL_ACCOUNTS_NAME.
    """
    creds = load_credentials()
    for c in creds:
        if c.get("name") == GLOBAL_ACCOUNTS_NAME:
            return c.get("accounts", {})
    return {}

def save_global_accounts_to_credentials(accounts: Dict) -> None:
    """
    Αποθηκεύει/ενημερώνει την εγγραφή GLOBAL_ACCOUNTS_NAME στο credentials.json
    με τα παρεχόμενα accounts mapping.
    """
    creds = load_credentials()
    found = False
    for i, c in enumerate(creds):
        if c.get("name") == GLOBAL_ACCOUNTS_NAME:
            creds[i]["accounts"] = accounts
            found = True
            break
    if not found:
        # προσθέτουμε ένα ειδικό credential αντικείμενο κρατώντας μόνο το accounts πεδίο
        creds.append({
            "name": GLOBAL_ACCOUNTS_NAME,
            "user": "",
            "key": "",
            "vat": "",
            "env": MYDATA_ENV,
            "accounts": accounts
        })
    save_credentials(creds)

# ---------------- Helpers ---------------- (most unchanged)
# -------------------------
# Robust JSON helpers
# -------------------------
def json_read(path, default=None):
    """
    Safe JSON read.
    - If file missing -> return default (default default: [] for lists, {} if you prefer)
    - If file empty/corrupt -> log and return default (do NOT raise).
    - Avoid recursive logging traps: use try/except carefully.
    """
    if default is None:
        default = []
    try:
        if not path or not os.path.exists(path):
            return default
        with open(path, "r", encoding="utf-8") as fh:
            txt = fh.read()
            if not txt or not txt.strip():
                return default
            # load
            return json.loads(txt)
    except Exception as e:
        # try a safe fallback: don't re-call json_read here (would recurse)
        try:
            log.error("json_read failed for %s: %s", path, str(e))
        except Exception:
            # if logging itself fails, fall back to print (rare)
            try:
                print("json_read failed for", path, str(e))
            except Exception:
                pass
        # return default to avoid crashing the app (caller should handle None/default)
        return default

def json_write(path, obj):
    """
    Atomic JSON write:
    - write to a tmp file, fsync, os.replace -> atomic replace
    - raises on failure so callers can handle/log
    """
    tmp = None
    try:
        dirp = os.path.dirname(path) or "."
        os.makedirs(dirp, exist_ok=True)
        text = json.dumps(obj, ensure_ascii=False, indent=2)
        # write to tmp file in same dir (for atomic replace)
        fd, tmp = tempfile.mkstemp(prefix=".tmp_json_", dir=dirp)
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            fh.write(text)
            fh.flush()
            try:
                os.fsync(fh.fileno())
            except Exception:
                # some file systems may not support fsync; ignore if fails
                pass
        # atomic replace
        os.replace(tmp, path)
        return True
    except Exception as e:
        try:
            log.exception("json_write failed for %s: %s", path, str(e))
        except Exception:
            print("json_write failed for", path, str(e))
        # cleanup tmp if present
        try:
            if tmp and os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        raise


def load_credentials():
    if os.path.exists(CREDENTIALS_FILE):
        with open(CREDENTIALS_FILE,'r',encoding='utf-8') as f:
            return json.load(f)
    return []

def save_credentials(credentials):
    with open(CREDENTIALS_FILE,'w',encoding='utf-8') as f:
        json.dump(credentials, f, ensure_ascii=False, indent=2)

def load_settings():
    if os.path.exists(SETTINGS_FILE):
        with open(SETTINGS_FILE,'r',encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_settings(settings):
    with open(SETTINGS_FILE,'w',encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

def get_active_credential():
    creds = load_credentials()
    for c in creds:
        if c.get('active'):
            return c
    return None

def add_credential(entry):
    creds = load_credentials()
    for c in creds:
        if c.get("name") == entry.get("name"):
            return False, "Credential with that name exists"
    creds.append(entry)
    save_credentials(creds)
    return True, ""

def update_credential(name, new_entry):
    creds = load_credentials()
    for i, c in enumerate(creds):
        if c.get("name") == name:
            creds[i] = new_entry
            save_credentials(creds)
            return True
    return False

def delete_credential(name):
    creds = load_credentials()
    new = [c for c in creds if c.get("name") != name]
    save_credentials(new)
    return True

def load_cache():
    data = json_read(CACHE_FILE)
    return data if isinstance(data, list) else []

def save_cache(docs):
    json_write(CACHE_FILE, docs)

def append_doc_to_cache(doc, aade_user=None, aade_key=None):
    docs = load_cache()
    try:
        sig = json.dumps(doc, sort_keys=True, ensure_ascii=False)
    except Exception:
        sig = json.dumps(str(doc), ensure_ascii=False)
    for d in docs:
        try:
            if json.dumps(d, sort_keys=True, ensure_ascii=False) == sig:
                return False
        except Exception:
            if str(d) == str(doc):
                return False
    docs.append(doc)
    save_cache(docs)
    return True

def save_summary_list(summary_list: List[Dict]):
    """Save summary_list to SUMMARY_FILE (overwrites)."""
    try:
        json_write(SUMMARY_FILE, summary_list)
    except Exception:
        log.exception("Could not write summary file")

# ---------------- small new helpers for per-customer files ----------------
def get_cred_by_name(name: str) -> Optional[Dict]:
    if not name:
        return None
    creds = load_credentials()
    return next((c for c in creds if c.get("name") == name), None)

# --- Excel path helper: πάντα με VAT + fiscal_year ---
def excel_path_for(vat: Optional[str] = None, cred_name: Optional[str] = None) -> str:
    """
    Return path to per-vat excel file: DATA_DIR/excel/<safe_vat>_<fiscal_year>_invoices.xlsx
    Fallback: if vat/cred_name missing return DEFAULT_EXCEL_FILE.
    """
    excel_dir = os.path.join(DATA_DIR, "excel")
    os.makedirs(excel_dir, exist_ok=True)

    # resolve fiscal year
    fy = None
    try:
        getter = globals().get("get_active_fiscal_year")
        if callable(getter):
            fy = getter()
    except Exception:
        fy = None
    if fy is None:
        from datetime import datetime
        fy = datetime.now().year

    if vat:
        try:
            safe_vat = secure_filename(str(vat))
        except Exception:
            safe_vat = str(vat)
        fname = f"{safe_vat}_{fy}_invoices.xlsx"
        return os.path.join(excel_dir, fname)

    if cred_name:
        try:
            safe_name = secure_filename(str(cred_name))
        except Exception:
            safe_name = str(cred_name)
        fname = f"{safe_name}_{fy}_invoices.xlsx"
        return os.path.join(excel_dir, fname)

    # fallback
    return DEFAULT_EXCEL_FILE


def _atomic_write(path, data_text):
    dirn = os.path.dirname(path)
    os.makedirs(dirn, exist_ok=True)
    fd, tmp_path = tempfile.mkstemp(dir=dirn, prefix=".tmp_")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(data_text)
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            try: os.remove(tmp_path)
            except Exception: pass

def _load_json(path):
    if not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def _save_json(path, obj):
    txt = json.dumps(obj, ensure_ascii=False, indent=2)
    _atomic_write(path, txt)

def _match_row_by_mark(df, mark):
    """
    Προσπαθεί να βρει γραμμή στο dataframe df όπου κάποια από τις πιθανές στήλες που
    αντιστοιχούν σε 'mark' ταιριάζει με την τιμή mark.
    Επιστρέφει (found_bool, row_dict_or_None, column_name_used_or_None)
    """
    possible_cols = [
        "mark", "MARK", "Mark",
        "invoice_id", "invoiceId", "id",
        "Αριθμός", "Αριθμός Μητρώου", "Αριθμός Μητρώου Παρ", "Α/Α",
        "αριθμος", "α/α"
    ]
    for col in possible_cols:
        if col in df.columns:
            # ανάγνωση ως string για ασφαλή σύγκριση
            matches = df[df[col].astype(str).str.strip() == str(mark).strip()]
            if not matches.empty:
                # επιστρέφουμε την πρώτη αντιστοιχία ως dict
                return True, matches.iloc[0].to_dict(), col
    # Γενικός fallback: ψάξε σε όλες τις στήλες που είναι string-like
    str_cols = [c for c in df.columns if df[c].dtype == 'object' or df[c].dtype == 'string']
    for c in str_cols:
        matches = df[df[c].astype(str).str.strip() == str(mark).strip()]
        if not matches.empty:
            return True, matches.iloc[0].to_dict(), c
    return False, None, None

def get_active_credential_from_session() -> Optional[Dict]:
    name = session.get("active_credential")
    return get_cred_by_name(name) if name else None

# NEW helper: epsilon per-vat path
def epsilon_file_path_for(vat: str) -> str:
    epsilon_dir = os.path.join(DATA_DIR, "epsilon")
    os.makedirs(epsilon_dir, exist_ok=True)
    return os.path.join(epsilon_dir, secure_filename(f"{vat}_epsilon_invoices.json"))

# New function: build epsilon from invoices.json (used when epsilon file missing)
def build_epsilon_from_invoices(vat: str) -> List[Dict]:
    """
    Δημιουργεί αρχική λίστα εγγραφών για το epsilon file (per-VAT) βασισμένη
    στο data/{vat}_invoices.json. Κάθε εγγραφή έχει:
      { "mark": "...", "AA": ..., "AFM": ..., "lines": [ {id, description, amount, vat, category:''}, ... ] }
    """
    invoices_file = get_customer_docs_file(vat)
    invoices = json_read(invoices_file) if os.path.exists(invoices_file) else []
    epsilon_list: List[Dict] = []

    def pick(src: dict, *keys, default=""):
        for k in keys:
            if k in src and src.get(k) not in (None, ""):
                return src.get(k)
        return default

    for doc in invoices:
        mark = str(pick(doc, "mark", "MARK", "Mark", default="")).strip()
        if not mark:
            mark = str(pick(doc, "identifier", "id", default="")).strip()
        if not mark:
            continue
        vat_doc = pick(doc, "AFM", "AFM_issuer", default="")
        aa = pick(doc, "AA", "aa", default="")
        raw_lines = doc.get("lines") or doc.get("Lines") or doc.get("Positions") or []
        prepared = []
        for idx, raw in enumerate(raw_lines):
            line_id = raw.get("id") or raw.get("line_id") or raw.get("LineId") or f"{mark}_l{idx}"
            description = pick(raw, "description", "desc", "Description", "name", "Name") or ""
            amount = pick(raw, "amount", "lineTotal", "net", "value", default="")
            vat_rate = pick(raw, "vat", "vatRate", "vatPercent", "vatAmount", default="")
            prepared.append({
                "id": line_id,
                "description": description,
                "amount": amount,
                "vat": vat_rate,
                "category": ""   # αρχικά κενό, user θα το συμπληρώσει
            })
        epsilon_list.append({
            "mark": mark,
            "AA": aa,
            "AFM": vat_doc,
            "lines": prepared
        })
    return epsilon_list

def load_epsilon_cache_for_vat(vat: str) -> List[Dict]:
    """
    Αν υπάρχει το per-vat epsilon αρχείο το διαβάζει, αλλιώς το χτίζει
    από το {vat}_invoices.json (με build_epsilon_from_invoices) και το αποθηκεύει.
    """
    path = epsilon_file_path_for(vat)
    if os.path.exists(path):
        try:
            return json_read(path)
        except Exception:
            log.exception("Could not read epsilon cache for %s", vat)
            return []
    # build from invoices if possible
    try:
        epsilon_list = build_epsilon_from_invoices(vat)
        if epsilon_list:
            try:
                json_write(path, epsilon_list)
                log.info("Built epsilon cache for %s from %s_invoices.json (%d entries)", vat, vat, len(epsilon_list))
            except Exception:
                log.exception("Could not write newly built epsilon cache for %s", vat)
        return epsilon_list
    except Exception:
        log.exception("Failed to build epsilon cache from invoices for %s", vat)
        return []

def save_epsilon_cache_for_vat(vat: str, data: List[Dict]):
    path = epsilon_file_path_for(vat)
    try:
        json_write(path, data)
    except Exception:
        log.exception("Could not write epsilon cache for %s", vat)

# στο app.py — κάτω από get_active_credential_from_session()
@app.context_processor
def inject_active_credential():
    """
    Εισάγει αυτόματα στα templates:
      - active_credential: όνομα credential ή None
      - active_credential_vat: ΑΦΜ του active credential (ή empty string)
      - app_settings: γενικές ρυθμίσεις εφαρμογής (φορτώνονται από SETTINGS_FILE)
    """
    active = get_active_credential_from_session()
    name = active.get("name") if active else None
    vat = active.get("vat") if active else ""
    # Load settings (fall back to empty dict)
    try:
        settings = load_settings() or {}
    except Exception:
        log.exception("Could not load settings for context processor")
        settings = {}
    return dict(active_credential=name, active_credential_vat=vat, app_settings=settings)


# ---------------- Validation helper ----------------
def normalize_input_date_to_iso(s: str):
    if not s:
        return None
    s = s.strip()
    try:
        dt = datetime.datetime.strptime(s, "%d/%m/%Y")
        return dt.date().isoformat()
    except ValueError:
        return None

# ---------------- safe render ----------------
def safe_render(template_name, **ctx):
    try:
        return render_template(template_name, **ctx)
    except Exception as e:
        tb = traceback.format_exc()
        log.error("Template rendering failed for %s: %s\n%s", template_name, str(e), tb)
        debug = os.getenv("FLASK_DEBUG", "0") == "1"
        body = "<h2>Template error</h2><p>" + escape(str(e)) + "</p>"
        if debug:
            body += "<pre>" + escape(tb) + "</pre>"
        return body

# ---------------- Routes ----------------
@app.route("/")
def home():
    return safe_render("nav.html", active_page="home")


@app.route('/get_fiscal_year', methods=['GET'])
def route_get_fiscal_year():
    """
    GET -> return current fiscal year if present:
    { exists: bool, fiscal_year: int|null }
    """
    y = get_active_fiscal_year()
    return jsonify(exists=(y is not None), fiscal_year=y), 200

@app.route('/set_fiscal_year', methods=['POST'])
def route_set_fiscal_year():
    """
    POST JSON or form: { fiscal_year: 2025 } or { year: 2025 }
    Returns JSON { success: bool, fiscal_year: int|null, message: str }
    """
    try:
        # accept JSON or form
        data = request.get_json(silent=True) or request.form or {}
        fy = data.get('fiscal_year') or data.get('year') or None
        if fy is None:
            return jsonify(success=False, message='Missing fiscal_year'), 400
        try:
            fy_int = int(fy)
        except Exception:
            return jsonify(success=False, message='Invalid fiscal_year'), 400

        ok = set_active_fiscal_year(fy_int)
        if not ok:
            return jsonify(success=False, message='Could not persist fiscal year'), 500
        return jsonify(success=True, fiscal_year=fy_int, message='Fiscal year updated'), 200
    except Exception:
        try:
            log.exception("Failed in set_fiscal_year")
        except Exception:
            pass
        return jsonify(success=False, message='Server error'), 500

# --- helpers για upsert / merge epsilon invoices -------------------------------
import uuid
from collections import OrderedDict

def _normalize_val(v):
    try:
        return "" if v is None else str(v).strip()
    except Exception:
        return ""

def _make_id_inv():
    return uuid.uuid4().hex

def _write_json_atomic(path, obj):
    """Atomic write (utf-8, indent=2)."""
    import tempfile
    dirn = os.path.dirname(path)
    os.makedirs(dirn, exist_ok=True)
    text = json.dumps(obj, ensure_ascii=False, indent=2)
    fd, tmp = tempfile.mkstemp(prefix=".tmp_json_", dir=dirn)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as fh:
            fh.write(text)
        os.replace(tmp, path)
    except Exception:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
        raise

def _upsert_epsilon_invoice(new_doc: dict):
    """
    Insert or update a detailed invoice record into the proper data/epsilon/*_epsilon_invoices.json.
    - Matches by mark (preferred) or AA.
    - If it finds an existing placeholder it replaces/merges it and preserves id_inv (or creates one).
    - If not found, it creates a per-vat file using AFM_issuer/AFM (if available) or 'unknown'.
    Returns (path_written, id_inv).
    """
    epsilon_dir = os.path.join(DATA_DIR, "epsilon")
    os.makedirs(epsilon_dir, exist_ok=True)

    mark = _normalize_val(new_doc.get("mark") or new_doc.get("MARK"))
    aa = _normalize_val(new_doc.get("AA") or new_doc.get("aa"))
    target_vat = _normalize_val(new_doc.get("AFM_issuer") or new_doc.get("AFM") or new_doc.get("issuer_vat"))

    # build candidate list: prefer per-vat file if target_vat present, then all others
    candidates = []
    if target_vat:
        candidates.append(os.path.join(epsilon_dir, f"{secure_filename(target_vat)}_epsilon_invoices.json"))
    for fname in sorted(os.listdir(epsilon_dir)):
        if not fname.endswith("_epsilon_invoices.json"):
            continue
        full = os.path.join(epsilon_dir, fname)
        if full not in candidates:
            candidates.append(full)

    # helper to extract list container & container_key info
    def _extract_items_and_container(raw):
        if isinstance(raw, list):
            return raw, True, None, raw  # items, is_root_list, key, root_container
        if isinstance(raw, dict):
            for k in ("items", "invoices", "data", "records"):
                if k in raw and isinstance(raw[k], list):
                    return raw[k], False, k, raw
            # maybe single invoice dict
            if any(k in raw for k in ("mark", "AA", "AFM", "AFM_issuer")):
                return [raw], False, None, raw
        return [], False, None, raw

    # try to find & replace/merge
    for path in candidates:
        try:
            with open(path, "r", encoding="utf-8") as fh:
                raw = json.load(fh)
        except Exception:
            # ignore unreadable files
            continue

        items, is_list, container_key, root_container = _extract_items_and_container(raw)
        if not items:
            continue

        changed = False
        for idx, it in enumerate(items):
            try:
                it_mark = _normalize_val(it.get("mark") or it.get("MARK"))
                it_aa = _normalize_val(it.get("AA") or it.get("aa"))
                # match exact or substring (covers small formatting diffs)
                matched = False
                if mark and it_mark and (mark == it_mark or mark in it_mark or it_mark in mark):
                    matched = True
                elif aa and it_aa and aa == it_aa:
                    matched = True

                if not matched:
                    continue

                # found matching existing item -> merge/replace
                existing = dict(it) if isinstance(it, dict) else {}
                # preserve existing id_inv if present, else take from new_doc, else create one
                id_inv = _normalize_val(existing.get("id_inv") or new_doc.get("id_inv") or new_doc.get("id") or "")
                if not id_inv:
                    id_inv = _make_id_inv()

                # merge: new_doc fields override existing ones; keep any remaining existing keys if not present in new_doc
                merged = dict(existing)
                merged.update(new_doc)  # new_doc wins
                merged["id_inv"] = id_inv

                # ensure id_inv appears BEFORE 'lines' key in JSON order
                new_ordered = OrderedDict()
                inserted = False
                for k, v in list(merged.items()):
                    if k == "lines" and not inserted:
                        new_ordered["id_inv"] = id_inv
                        inserted = True
                    new_ordered[k] = v
                if not inserted:
                    # append id_inv at start (or end - we put at start to be "before lines")
                    od = OrderedDict()
                    od["id_inv"] = id_inv
                    for k,v in new_ordered.items():
                        od[k] = v
                    new_ordered = od

                # replace the item in the list
                items[idx] = dict(new_ordered)
                changed = True
                # write back updated container
                if changed:
                    if is_list:
                        to_write = items
                    else:
                        if container_key:
                            root_container[container_key] = items
                            to_write = root_container
                        else:
                            # single dict root replaced by this new item
                            to_write = items[0]
                    _write_json_atomic(path, to_write)
                    try:
                        log.info("_upsert_epsilon_invoice: updated %s (mark=%s AA=%s id_inv=%s)", path, mark, aa, id_inv)
                    except Exception:
                        pass
                    return path, id_inv

            except Exception:
                continue

    # not found anywhere -> create per-vat file (prefer AFM_issuer/AFM), else 'unknown'
    safe_vat = secure_filename(target_vat) if target_vat else "unknown"
    new_path = os.path.join(epsilon_dir, f"{safe_vat}_epsilon_invoices.json")
    id_inv = _normalize_val(new_doc.get("id_inv") or new_doc.get("id") or "")
    if not id_inv:
        id_inv = _make_id_inv()

    # ensure id_inv before lines
    merged = dict(new_doc)
    merged["id_inv"] = id_inv
    new_ordered = OrderedDict()
    inserted = False
    for k, v in list(merged.items()):
        if k == "lines" and not inserted:
            new_ordered["id_inv"] = id_inv
            inserted = True
        new_ordered[k] = v
    if not inserted:
        od = OrderedDict()
        od["id_inv"] = id_inv
        for k,v in new_ordered.items():
            od[k] = v
        new_ordered = od

    # write single-element list to new_path (or append if file exists and is a list)
    try:
        if os.path.exists(new_path):
            # if file exists and is a list, load & append
            with open(new_path, "r", encoding="utf-8") as fh:
                raw = json.load(fh)
            if isinstance(raw, list):
                raw.append(dict(new_ordered))
                _write_json_atomic(new_path, raw)
            elif isinstance(raw, dict):
                # try to append into detected container key if possible, else rewrite as list
                appended = False
                for k in ("items","invoices","data","records"):
                    if k in raw and isinstance(raw[k], list):
                        raw[k].append(dict(new_ordered))
                        _write_json_atomic(new_path, raw)
                        appended = True
                        break
                if not appended:
                    # convert to list form
                    _write_json_atomic(new_path, [dict(new_ordered)])
            else:
                _write_json_atomic(new_path, [dict(new_ordered)])
        else:
            _write_json_atomic(new_path, [dict(new_ordered)])
        try:
            log.info("_upsert_epsilon_invoice: created %s (mark=%s AA=%s id_inv=%s)", new_path, mark, aa, id_inv)
        except Exception:
            pass
        return new_path, id_inv
    except Exception:
        try:
            log.exception("_upsert_epsilon_invoice: write failed for %s", new_path)
        except Exception:
            pass
        return "", ""
# -------------------------------------------------------------------------------

# ---------- validation helpers ----------
def parse_date_str_to_utc(date_str):
    """
    Try parse a date string (ISO-ish or dd/mm/yyyy or yyyy-mm-dd).
    Returns a datetime in UTC (naive or tz aware converted to UTC) or None.
    """
    if not date_str:
        return None
    # try ISO first
    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            dt = _dt.strptime(date_str, fmt)
            # naive -> assume local? we'll treat naive as UTC to be strict
            if dt.tzinfo is None:
                # treat as UTC for server-side canonicalization
                dt = dt.replace(tzinfo=timezone.utc)
            else:
                dt = dt.astimezone(timezone.utc)
            return dt
        except Exception:
            continue
    # last resort: try dateutil if available
    try:
        from dateutil import parser as _parser
        dt = _parser.parse(date_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt
    except Exception:
        return None

def is_date_within_fiscal_year(dt_utc, fiscal_year):
    """
    dt_utc: datetime with tzinfo=UTC
    fiscal_year: int (e.g. 2025)
    Assumes fiscal year = calendar year (Jan 1 - Dec 31).
    If your fiscal year differs, adapt start/end calculation here.
    """
    if not dt_utc or fiscal_year is None:
        return False
    start = _dt(fiscal_year, 1, 1, 0, 0, 0, tzinfo=timezone.utc)
    end = _dt(fiscal_year, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
    return start <= dt_utc <= end

def validate_date_field_against_active_fiscal(date_str, field_name='date'):
    """
    Centralized server-side check. Returns (ok:bool, message:str).
    Use this in any route that receives date strings from user.
    """
    fiscal_year = get_active_fiscal_year()
    if fiscal_year is None:
        return True, "No active fiscal year set"  # if no selection, don't block
    dt = parse_date_str_to_utc(date_str)
    if dt is None:
        return False, f"Δεν αναγνώστηκε σωστά η ημερομηνία στο πεδίο {field_name}."
    if not is_date_within_fiscal_year(dt, fiscal_year):
        # produce message with allowed year info
        return False, f"Η επιλεγμένη χρήση είναι {fiscal_year}. Η ημερομηνία στο πεδίο {field_name} ({date_str}) δεν ανήκει στη χρήση αυτή."
    return True, "OK"

@app.route('/api/repeat_entry/get', methods=['GET'])
def api_repeat_entry_get():
    """
    Επιστρέφει στοιχεία repeat_entry + expense_tags και -όταν είναι διαθέσιμο-
    το afm/vat του ενεργού πελάτη ώστε το frontend να το χρησιμοποιεί.
    Παράμετρος query: ?vat=... (προαιρετικό)
    """
    try:
        creds = read_credentials_list()
    except Exception:
        creds = None

    # default empty response structure
    base_resp = {"ok": True, "repeat_entry": {"enabled": False, "mapping": {}}, "expense_tags": []}

    # try to get vat from query param or from session active credential
    vat_param = request.args.get('vat') or None
    session_cred = (get_active_credential_from_session() or {}) if 'get_active_credential_from_session' in globals() else {}
    vat_from_session = session_cred.get('vat') if isinstance(session_cred, dict) else None
    vat = vat_param or vat_from_session

    if not creds:
        # still include any afm/vat from session if present
        afm_from_session = ""
        try:
            # try common keys
            if isinstance(session_cred, dict):
                afm_from_session = (session_cred.get('afm') or session_cred.get('vat') or "") or ""
            afm_from_session = str(afm_from_session).strip() if afm_from_session else ""
        except Exception:
            afm_from_session = ""
        resp = dict(base_resp)
        if afm_from_session:
            resp['afm'] = afm_from_session
            resp['vat'] = afm_from_session
        return jsonify(resp)

    # find index of active client if possible
    idx = find_active_client_index(creds, vat=vat) if 'find_active_client_index' in globals() else None
    if idx is None:
        # not found — still try to return session info if available
        afm_guess = ""
        try:
            if isinstance(session_cred, dict):
                afm_guess = (session_cred.get('afm') or session_cred.get('vat') or "") or ""
            afm_guess = str(afm_guess).strip() if afm_guess else ""
        except Exception:
            afm_guess = ""

        resp = dict(base_resp)
        # if we can extract expense_tags from session_cred, include them
        try:
            raw_tags = (session_cred.get('expense_tags') if isinstance(session_cred, dict) else None) or []
            if isinstance(raw_tags, str):
                expense_tags = [t.strip() for t in raw_tags.split(',') if t.strip()]
            elif isinstance(raw_tags, list):
                expense_tags = [str(t) for t in raw_tags]
            else:
                expense_tags = []
            if expense_tags:
                resp['expense_tags'] = expense_tags
        except Exception:
            pass

        if afm_guess:
            resp['afm'] = afm_guess
            resp['vat'] = afm_guess
        return jsonify(resp)

    # we have creds and an index -> build response
    client_rec = creds[idx] if idx is not None and idx < len(creds) else None
    repeat = (client_rec.get('repeat_entry') if isinstance(client_rec, dict) else {}) or {"enabled": False, "mapping": {}}

    # normalize expense_tags
    raw_tags = (client_rec.get('expense_tags') if isinstance(client_rec, dict) else None) or []
    if isinstance(raw_tags, str):
        expense_tags = [t.strip() for t in raw_tags.split(',') if t.strip()]
    elif isinstance(raw_tags, list):
        expense_tags = [str(t) for t in raw_tags]
    else:
        expense_tags = []

    # try to extract afm/vat from client_rec (many possible key names)
    afm_found = ""
    vat_found = ""
    try:
        if isinstance(client_rec, dict):
            # check common keys
            for k in ("afm", "AFM", "vat", "VAT", "vat_number", "vatNumber"):
                v = client_rec.get(k)
                if v:
                    val = str(v).strip()
                    if not afm_found and (k.lower().startswith('afm') or (len(val) == 9 and val.isdigit())):
                        afm_found = val
                    if not vat_found and (k.lower().startswith('vat') or (len(val) >= 8)):
                        vat_found = val
            # nested structures
            for nested in ("active_client", "client", "credential"):
                if not afm_found and isinstance(client_rec.get(nested), dict):
                    nc = client_rec.get(nested)
                    for k in ("afm", "AFM", "vat", "VAT", "vat_number"):
                        if k in nc and nc[k]:
                            afm_found = afm_found or str(nc[k]).strip()
                            vat_found = vat_found or str(nc[k]).strip()
            # as fallback use name fields (not AFM, but useful info)
            client_name = client_rec.get('client_name') or client_rec.get('name') or client_rec.get('company') or client_rec.get('client') or ""
        else:
            client_name = ""
    except Exception:
        afm_found = afm_found or ""
        vat_found = vat_found or ""
        client_name = client_name if 'client_name' in locals() else ""

    # fallback to session credential values if nothing found
    try:
        if not afm_found and isinstance(session_cred, dict):
            afm_found = (session_cred.get('afm') or session_cred.get('vat') or "") or afm_found
        if not vat_found and isinstance(session_cred, dict):
            vat_found = (session_cred.get('vat') or session_cred.get('afm') or "") or vat_found
    except Exception:
        pass

    afm_found = (str(afm_found).strip() or "")
    vat_found = (str(vat_found).strip() or "")

    resp = {
        "ok": True,
        "repeat_entry": repeat,
        "expense_tags": expense_tags
    }
    # include descriptive client object (don't leak secrets) - include some fields if present
    try:
        client_summary = {}
        if isinstance(client_rec, dict):
            for fld in ("client_name", "name", "company"):
                if fld in client_rec and client_rec.get(fld):
                    client_summary["name"] = client_rec.get(fld)
                    break
            # include provided vat/afm if present
            if afm_found:
                client_summary["afm"] = afm_found
            elif vat_found:
                client_summary["vat"] = vat_found
            # include original entry for debugging only when debug enabled (optional)
            if app.debug:
                client_summary["_raw"] = client_rec
        if client_summary:
            resp["client"] = client_summary
    except Exception:
        pass

    if afm_found:
        resp['afm'] = afm_found
    elif vat_found:
        resp['vat'] = vat_found

    # also return the vat param echoed (helpful for debugging)
    if vat_param:
        resp['_queried_vat'] = vat_param

    return jsonify(resp)



@app.route('/api/repeat_entry/save', methods=['POST'])
def api_repeat_entry_save():
    payload = request.get_json(silent=True) or {}
    enabled = bool(payload.get('enabled', False))
    mapping = payload.get('mapping', {}) or {}

    # sanitize + normalize mapping keys and values
    normalized_mapping = {}
    for k, v in mapping.items():
        nk = normalize_vat_key(k)
        if not nk:
            # allow storing __default specially (leave as-is)
            if str(k).strip() == '__default' and v:
                normalized_mapping['__default'] = str(v).strip()
            continue
        normalized_mapping[nk] = str(v).strip()

    creds = read_credentials_list()
    vat = payload.get('vat') or (get_active_credential_from_session() or {}).get('vat')
    idx = find_active_client_index(creds, vat=vat)
    if idx is None:
        return jsonify({"ok": False, "error": "No credentials found"}), 400

    creds[idx].setdefault('repeat_entry', {})
    creds[idx]['repeat_entry']['enabled'] = enabled
    creds[idx]['repeat_entry']['mapping'] = normalized_mapping
    creds[idx]['repeat_entry']['updated_at'] = datetime.datetime.utcnow().isoformat() + 'Z'
    creds[idx]['repeat_entry']['user'] = (get_active_credential_from_session() or {}).get('user') or 'unknown'

    try:
        write_credentials_list(creds)
        # optionally: return expense_tags too, if you earlier wanted that
        return jsonify({"ok": True, "repeat_entry": creds[idx]['repeat_entry']})
    except Exception as e:
        log.exception("Failed saving repeat_entry")
        return jsonify({"ok": False, "error": str(e)}), 500



@app.route('/credentials/add', methods=['POST'])
def credentials_add():
    credentials = load_credentials()
    name = request.form.get('name')
    vat = request.form.get('vat')
    user = request.form.get('user')
    key = request.form.get('key')
    book_category = request.form.get('book_category','Β')
    fpa_applicable = bool(request.form.get('fpa_applicable'))
    expense_tags = request.form.getlist('expense_tags')
    apodeixakia_type = request.form.get('apodeixakia_type', '').strip()
    apodeixakia_supplier = request.form.get('apodeixakia_supplier', '').strip()
    new_cred = {
        'name': name,
        'vat': vat,
        'user': user,
        'key': key,
        'book_category': book_category,
        'fpa_applicable': fpa_applicable,
        'expense_tags': expense_tags,
        'apodeixakia_type': apodeixakia_type,
        'apodeixakia_supplier': apodeixakia_supplier,
        'active': False
    }
    credentials.append(new_cred)
    save_credentials(credentials)
    # <-- added flash to ensure message appears if this route is used
    flash("Saved", "success")
    return redirect(url_for('credentials'))




@app.route('/credentials/delete/<name>', methods=['POST'])
def credentials_delete_post(name):
    credentials = load_credentials()
    credentials = [c for c in credentials if c['name'] != name]
    save_credentials(credentials)
    # <-- added flash to ensure message appears if this route is used
    flash(f"Credential {name} διαγράφηκε.", "success")
    return redirect(url_for('credentials'))

@app.route('/credentials/set_active', methods=['POST'])
def credentials_set_active():
    active_name = request.form.get('active_name')
    credentials = load_credentials()
    for c in credentials:
        c['active'] = (c['name'] == active_name)
    save_credentials(credentials)
    # <-- added flash in case this route is used
    flash(f"Active credential set to {active_name}", "success")
    return redirect(url_for('credentials'))

@app.route('/credentials/save_settings', methods=['POST'])
def credentials_save_settings():
    data = request.get_json()
    if data:
        save_settings(data)
        return jsonify({'status':'ok'})
    return jsonify({'status':'error'}), 400

@app.route('/api/save_receipt', methods=['POST'])
def api_save_receipt():
    """
    Expects JSON { "summary": {...} }
    Writes ONLY to epsilon_invoices.json and to Excel.
    Returns JSON { ok: True, duplicate: Bool, message: str }
    """
    try:
        payload = request.get_json(silent=True) or {}
        summary = payload.get('summary') or {}
        if not summary:
            return jsonify(ok=False, error='no summary provided'), 400

        rec = _normalize_receipt_summary(summary)

        # Acquire JSON lock and check duplicates & append
        json_lock = FileLock(EPSILON_JSON_LOCK, timeout=10)
        with json_lock:
            epsilon_list = _read_epsilon_json()
            if _is_duplicate_in_epsilon(epsilon_list, rec):
                return jsonify(ok=True, duplicate=True, message='duplicate entry'), 200
            # append and write
            epsilon_list.append(rec)
            try:
                _write_epsilon_json(epsilon_list)
            except Exception as e:
                current_app.logger.exception("failed writing epsilon json: %s", e)
                return jsonify(ok=False, error='write epsilon json failed'), 500

        # Append to Excel with separate lock
        excel_lock = FileLock(EXCEL_FILE_LOCK, timeout=10)
        with excel_lock:
            ok_excel = _append_to_excel(rec)
            if not ok_excel:
                current_app.logger.warning("Excel append failed for record: %s", rec)
                # we proceed since epsilon json already saved - but inform client
                return jsonify(ok=True, duplicate=False, warning='excel_append_failed', message='saved to epsilon json but excel append failed'), 200

        return jsonify(ok=True, duplicate=False, message='saved'), 200

    except Exception as e:
        current_app.logger.exception("api_save_receipt error: %s", e)
        return jsonify(ok=False, error=str(e)), 500

@app.route('/upload_client_db', methods=['POST'])
def upload_client_db():
    """
    Accept multipart/form-data with field 'client_file' and save it into DATA_DIR
    as client_db{.ext}. Existing client_db* files are moved to a single backup
    (previous backups removed), and client_db.meta.json is written.
    Returns JSON { success: bool, message: str, missing_columns: [...], detected_columns: [...],
                   uploaded_at: str, total_rows: int, new_clients: int, existing_clients: int }
    """
    try:
        os.makedirs(DATA_DIR, exist_ok=True)

        if 'client_file' not in request.files:
            return jsonify(success=False, message='Δεν βρέθηκε το πεδίο client_file στο αίτημα.'), 400

        f = request.files['client_file']
        if not f or not getattr(f, 'filename', '').strip():
            return jsonify(success=False, message='Δεν επιλέχθηκε αρχείο.'), 400

        filename = secure_filename(f.filename)
        base, ext = os.path.splitext(filename)
        ext = ext.lower()
        if ext not in ALLOWED_CLIENT_EXT:
            return jsonify(success=False, message='Μη επιτρεπτή επέκταση. Χρήση .xlsx, .xls ή .csv'), 400

        # --- Έλεγχος headers ---
        try:
            stream = f.stream
            success, headers, err = _extract_headers_from_upload(stream, ext)
            if not success:
                return jsonify(success=False, message=err or 'Αποτυχία ανάγνωσης αρχείου για έλεγχο headers.'), 400

            headers_set = {str(h).strip() for h in headers}
            missing = sorted(list(REQUIRED_CLIENT_COLUMNS - headers_set))
            if missing:
                return jsonify(success=False,
                               message='Λείπουν υποχρεωτικές στήλες.',
                               missing_columns=missing,
                               detected_columns=sorted(list(headers_set))), 400
        except Exception as e:
            log.exception("Error while extracting headers from uploaded client_file")
            return jsonify(success=False, message=f'Σφάλμα κατά τον έλεγχο των στηλών: {e}'), 500

        # --- Ανάγνωση client_file για επεξεργασία πελατών ---
        try:
            f.stream.seek(0)
            if ext in ['.xls', '.xlsx']:
                df = pd.read_excel(f.stream, dtype=str)
            else:
                df = pd.read_csv(f.stream, dtype=str)
            df.fillna('', inplace=True)
        except Exception as e:
            log.exception("Failed to read uploaded client_file")
            return jsonify(success=False, message=f'Σφάλμα κατά την ανάγνωση του αρχείου: {e}'), 500

        total_rows = len(df)
        existing_clients_set = get_existing_client_ids()
        new_clients_set = set()
        already_existing_set = set()

        for afm in df.get("ΑΦΜ", []):
            afm_str = str(afm).strip()
            if afm_str:
                if afm_str in existing_clients_set:
                    already_existing_set.add(afm_str)
                else:
                    new_clients_set.add(afm_str)

        # --- Backup: keep only one backup ---
        dest_name = f'client_db{ext}'
        dest_path = os.path.join(DATA_DIR, dest_name)

        try:
            # 1) remove any previous backups (files that contain ".bak." or end with ".bak")
            for existing in os.listdir(DATA_DIR):
                if not existing.startswith('client_db'):
                    continue
                # consider files like client_db.xlsx.bak.20250101T000000Z or client_db.bak
                if '.bak.' in existing or existing.endswith('.bak'):
                    try:
                        os.remove(os.path.join(DATA_DIR, existing))
                        log.info("Removed old client_db backup: %s", existing)
                    except Exception:
                        log.exception("Failed to remove old backup %s (continuing)", existing)

            # 2) move current client_db.* (if any) to a new single backup (timestamped)
            for existing in os.listdir(DATA_DIR):
                if existing.startswith('client_db'):
                    existing_ext = os.path.splitext(existing)[1].lower()
                    if existing_ext in ALLOWED_CLIENT_EXT:
                        existing_path = os.path.join(DATA_DIR, existing)
                        ts = _dt.utcnow().strftime('%Y%m%dT%H%M%SZ')
                        backup_name = f"{existing}.bak.{ts}"
                        backup_path = os.path.join(DATA_DIR, backup_name)
                        try:
                            os.rename(existing_path, backup_path)
                            log.info("Backed up previous client_db: %s -> %s", existing, backup_name)
                        except Exception:
                            log.exception("Failed to backup previous client_db %s (continuing)", existing)
        except Exception:
            log.exception("Failed while rotating client_db backups (continuing)")

        # --- Save uploaded file to destination path ---
        try:
            f.stream.seek(0)
            f.save(dest_path)
        except Exception:
            log.exception("Failed to save uploaded client_db to %s", dest_path)
            return jsonify(success=False, message='Σφάλμα κατά την αποθήκευση του αρχείου.'), 500

        # --- Meta info ---
        try:
            uploaded_at_iso = _dt.utcnow().replace(microsecond=0).isoformat() + 'Z'
            write_client_meta(filename, uploaded_at_iso)
        except Exception:
            log.exception("Failed to write client_db metadata (continuing)")

        return jsonify(
            success=True,
            message=f'Αποθηκεύτηκε: {dest_name}',
            filename=filename,
            uploaded_at=uploaded_at_iso,
            detected_columns=sorted(list(headers_set)),
            total_rows=total_rows,
            new_clients=len(new_clients_set),
            existing_clients=len(already_existing_set)
        ), 200

    except Exception:
        log.exception("Unhandled exception in upload_client_db")
        return jsonify(success=False, message='Εσωτερικό σφάλμα server.'), 500




# --- client_db_info route ---
@app.route('/client_db_info', methods=['GET'])
def client_db_info():
    """
    Return JSON with metadata about current client_db (if exists).
    { exists: bool, filename: str|null, uploaded_at: str|null, total_rows: int, new_rows: int, updated_rows: int }
    """
    try:
        meta = read_client_meta()
        counts = {'total_rows': 0, 'new_rows': 0, 'updated_rows': 0}

        if meta:
            # αν υπάρχει client_db, διαβάζουμε για μέτρηση
            p = os.path.join(DATA_DIR, f"client_db{os.path.splitext(meta.get('filename',''))[1]}")
            if os.path.exists(p):
                try:
                    ext = os.path.splitext(p)[1].lower()
                    if ext in ['.xls', '.xlsx']:
                        df = pd.read_excel(p, dtype=str)
                    else:
                        df = pd.read_csv(p, dtype=str)
                    df.fillna('', inplace=True)
                    total_rows = len(df)
                    existing_clients_set = set(get_existing_client_ids())
                    new_rows = updated_rows = 0
                    for idx, row in df.iterrows():
                        client_id = str(row.get("ΑΦΜ", "")).strip()
                        if not client_id:
                            continue
                        if client_id in existing_clients_set:
                            updated_rows += 1
                        else:
                            new_rows += 1
                            existing_clients_set.add(client_id)
                    counts.update({'total_rows': total_rows, 'new_rows': new_rows, 'updated_rows': updated_rows})
                except Exception:
                    log.exception("Failed to count client_db rows")

            return jsonify(exists=True, filename=meta.get('filename'),
                           uploaded_at=meta.get('uploaded_at'),
                           **counts), 200

        # fallback: if any client_db.* exists but no meta file
        for existing in os.listdir(DATA_DIR):
            if existing.startswith('client_db') and os.path.splitext(existing)[1].lower() in ALLOWED_CLIENT_EXT:
                p = os.path.join(DATA_DIR, existing)
                try:
                    mtime = _dt.utcfromtimestamp(os.path.getmtime(p)).replace(microsecond=0).isoformat() + 'Z'
                    counts.update({'total_rows': 0, 'new_rows': 0, 'updated_rows': 0})
                    return jsonify(exists=True, filename=existing, uploaded_at=mtime, **counts), 200
                except Exception:
                    continue

        return jsonify(exists=False, filename=None, uploaded_at=None, **counts), 200
    except Exception:
        log.exception("Failed to read client_db info")
        return jsonify(exists=False, filename=None, uploaded_at=None, total_rows=0, new_rows=0, updated_rows=0), 500
# ---------- end client_db helpers + routes ----------

# credentials CRUD (unchanged behaviour) but pass active credential to template
@app.route("/credentials", methods=["GET", "POST"])
def credentials():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        user = request.form.get("user", "").strip()
        key = request.form.get("key", "").strip()
        env = MYDATA_ENV
        vat = request.form.get("vat", "").strip()

        # Νέα πεδία
        book_category = request.form.get("book_category", "Β").strip() or "Β"
        fpa_applicable = True if request.form.get("fpa_applicable") in ("on", "true", "1") else False
        expense_tags = request.form.getlist("expense_tags") or []
        apodeixakia_type = request.form.get("apodeixakia_type", "").strip()   # "afm" or "supplier"
        apodeixakia_supplier = request.form.get("apodeixakia_supplier", "").strip()

        if not name:
            flash("Name required", "error")
        else:
            entry = {
                "name": name,
                "user": user,
                "key": key,
                "env": env,
                "vat": vat,
                # προσθήκη νέων πεδίων
                "book_category": book_category,
                "fpa_applicable": fpa_applicable,
                "expense_tags": expense_tags,
                "apodeixakia_type": apodeixakia_type,
                "apodeixakia_supplier": apodeixakia_supplier
            }
            ok, err = add_credential(entry)
            if ok:
                flash("Saved", "success")
            else:
                flash(err or "Could not save", "error")
        return redirect(url_for("credentials"))

    # GET: φορτώνουμε τα credentials και αφήνουμε το context_processor να περάσει το active credential/ΑΦΜ
    creds = load_credentials()
    return safe_render("credentials_list.html", credentials=creds, active_page="credentials")


@app.route("/credentials/edit/<name>", methods=["GET", "POST"])
def credentials_edit(name):
    creds = load_credentials()
    credential = next((c for c in creds if c.get("name") == name), None)
    if not credential:
        flash("Credential not found", "error")
        return redirect(url_for("credentials"))

    if request.method == "POST":
        new_name = (request.form.get("name") or "").strip()
        user = (request.form.get("user") or "").strip()
        key = (request.form.get("key") or "").strip()
        env = (request.form.get("env") or MYDATA_ENV).strip()
        vat = (request.form.get("vat") or "").strip()

        # Νέα πεδία από τη φόρμα επεξεργασίας
        book_category = request.form.get("book_category", "Β").strip() or "Β"
        fpa_applicable = True if request.form.get("fpa_applicable") in ("on", "true", "1") else False
        expense_tags = request.form.getlist("expense_tags") or []

        # Νέα πεδία (apodeixakia) - fallback στις υπάρχουσες τιμές αν δεν παρέχονται
        apodeixakia_type = request.form.get("apodeixakia_type")
        if apodeixakia_type is None:
            apodeixakia_type = credential.get("apodeixakia_type", "")
        else:
            apodeixakia_type = apodeixakia_type.strip()

        apodeixakia_supplier = request.form.get("apodeixakia_supplier")
        if apodeixakia_supplier is None:
            apodeixakia_supplier = credential.get("apodeixakia_supplier", "")
        else:
            apodeixakia_supplier = apodeixakia_supplier.strip()

        if not new_name:
            flash("Name required", "error")
            return redirect(url_for("credentials_edit", name=name))

        if new_name != name and any(c.get("name") == new_name for c in creds):
            flash("Another credential with that name already exists", "error")
            return redirect(url_for("credentials_edit", name=name))

        new_entry = {
            "name": new_name,
            "user": user,
            "key": key,
            "env": env,
            "vat": vat,
            # Αποθηκεύουμε επίσης τα νέα πεδία
            "book_category": book_category,
            "fpa_applicable": fpa_applicable,
            "expense_tags": expense_tags,
            # Αποθηκεύουμε και τα apodeixakia πεδία
            "apodeixakia_type": apodeixakia_type,
            "apodeixakia_supplier": apodeixakia_supplier
        }

        updated = False
        for i, c in enumerate(creds):
            if c.get("name") == name:
                creds[i] = new_entry
                updated = True
                break
        if not updated:
            creds.append(new_entry)

        save_credentials(creds)

        # Αν το credential που επεξεργάστηκε ήταν ενεργό — ενημέρωσε session
        if session.get("active_credential") == name:
            session["active_credential"] = new_name
            flash(f"Active credential updated to '{new_name}'", "success")
        else:
            flash(f"Credential '{new_name}' updated successfully", "success")

        return redirect(url_for("credentials"))

    # GET -> εμφανίζουμε τη φόρμα επεξεργασίας
    # Προσθέτουμε flash πληροφορία (παραμένει ως έχει)
    flash(f"Editing credential: {credential.get('name')}", "info")
    return safe_render(
        "credentials_edit.html",
        credential=credential,
        active_page="credentials"
    )


@app.route("/credentials/delete/<name>", methods=["POST"])
def credentials_delete(name):
    creds = load_credentials()
    credential = next((c for c in creds if c.get("name") == name), None)

    if not credential:
        # Αν είναι AJAX, επιστρέφουμε JSON
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'status': 'error', 'error': 'not found'}), 404
        flash(f"Credential '{name}' not found", "error")
        return redirect(url_for("credentials"))

    # Αφαίρεση credential
    creds = [c for c in creds if c.get("name") != name]
    save_credentials(creds)

    # Αν ήταν ενεργό, καθαρίζουμε session
    was_active = False
    if session.get("active_credential") == name:
        session.pop("active_credential", None)
        was_active = True

    # Αν request από AJAX, επιστρέφουμε JSON ώστε το frontend fetch να το χειριστεί
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'status': 'ok', 'deleted': name, 'was_active': was_active})

    # Διαφορετικά κάνουμε τα flash + redirect όπως παλιά
    if was_active:
        flash(f"Active credential '{name}' διαγράφηκε και αφαιρέθηκε από τα ενεργά.", "success")
    else:
        flash(f"Credential '{name}' διαγράφηκε.", "success")

    return redirect(url_for("credentials"))


# New route: set active credential
@app.route("/set_active", methods=["POST"])
def set_active_credential():
    name = request.form.get("active_name")
    if not name:
        flash("No credential selected", "error")
    else:
        cred = get_cred_by_name(name)
        if not cred:
            flash("Credential not found", "error")
        else:
            session["active_credential"] = name
            flash(f"Active credential set to {name}", "success")
    return redirect(url_for("credentials"))


# ---------------- Fetch page ----------------
# ---------------- Helpers for per-customer JSON & summary ----------------
def append_doc_to_customer_file(doc, vat):
    """
    Add a doc to per-customer JSON file, avoiding duplicates.
    Filename: data/{VAT}_invoices.json
    """
    if not vat:
        return False
    customer_file = os.path.join(DATA_DIR, f"{vat}_invoices.json")
    cache = json_read(customer_file)
    sig = json.dumps(doc, sort_keys=True, ensure_ascii=False)
    for d in cache:
        try:
            if json.dumps(d, sort_keys=True, ensure_ascii=False) == sig:
                return False
        except Exception:
            if str(d) == str(doc):
                return False
    cache.append(doc)
    json_write(customer_file, cache)
    return True

def append_summary_to_customer_file(summary, vat):
    """
    Save summary for a customer into per-customer summary JSON
    Filename: data/{VAT}_summary.json
    """
    if not vat:
        return False
    summary_file = os.path.join(DATA_DIR, f"{vat}_summary.json")
    summaries = json_read(summary_file)
    # avoid duplicate by MARK
    mark = str(summary.get("mark", "")).strip()
    if any(str(s.get("mark", "")).strip() == mark for s in summaries):
        return False
    summaries.append(summary)
    json_write(summary_file, summaries)
    return True

def get_customer_summary_file(vat):
    return os.path.join(DATA_DIR, f"{vat}_summary.json")

def get_customer_docs_file(vat):
    return os.path.join(DATA_DIR, f"{vat}_invoices.json")



# ---------------- Fetch page (updated with per-customer summary) ----------------

@app.route("/fetch", methods=["GET", "POST"])
def fetch():
    message = None
    error = None
    preview = []
    creds = load_credentials()
    active_cred = get_active_credential_from_session()
    active_name = active_cred.get("name") if active_cred else None

    if request.method == "POST":
        date_from_raw = request.form.get("date_from", "").strip()
        date_to_raw = request.form.get("date_to", "").strip()
        date_from_iso = normalize_input_date_to_iso(date_from_raw)
        date_to_iso = normalize_input_date_to_iso(date_to_raw)

        if not date_from_iso or not date_to_iso:
            error = "Παρακαλώ συμπλήρωσε έγκυρες ημερομηνίες (dd/mm/YYYY)."
            return safe_render("fetch.html", credentials=creds, message=message,
                               error=error, preview=preview, active_page="fetch",
                               active_credential=active_name)

        d1 = datetime.datetime.fromisoformat(date_from_iso).strftime("%d/%m/%Y")
        d2 = datetime.datetime.fromisoformat(date_to_iso).strftime("%d/%m/%Y")

        selected = request.form.get("use_credential") or session.get("active_credential") or ""
        vat = request.form.get("vat_number", "").strip()
        aade_user = AADE_USER_ENV
        aade_key = AADE_KEY_ENV
        if selected:
            c = next((x for x in creds if x.get("name") == selected), None)
            if c:
                aade_user = c.get("user") or aade_user
                aade_key = c.get("key") or aade_key
                vat = vat or c.get("vat", "")
                session["active_credential"] = c.get("name")

        if not aade_user or not aade_key:
            error = "Δεν υπάρχουν αποθηκευμένα credentials για την κλήση."
            return safe_render("fetch.html", credentials=creds, message=message,
                               error=error, preview=preview, active_page="fetch",
                               active_credential=active_name)

        try:
            all_rows, summary_list = request_docs(
                date_from=d1,
                date_to=d2,
                mark="000000000000000",
                aade_user=aade_user,
                aade_key=aade_key,
                debug=True,
                save_excel=False
            )

            added_docs = 0
            added_summaries = 0
            for d in all_rows:
                if vat:
                    d["AFM_counterpart"] = vat  # προσθέτουμε AFM
                if append_doc_to_customer_file(d, vat):
                    added_docs += 1

            for s in summary_list:
                if append_summary_to_customer_file(s, vat):
                    added_summaries += 1

            message = (f"Fetched {len(all_rows)} items, newly saved for VAT {vat}: "
                       f"{added_docs} docs, {added_summaries} summaries.")

            preview = all_rows[:40]

        except Exception as e:
            log.exception("Fetch error")
            error = f"Σφάλμα λήψης: {str(e)[:400]}"

    return safe_render("fetch.html", credentials=creds, message=message,
                       error=error, preview=preview, active_page="fetch",
                       active_credential=active_name)


@app.route("/credentials/get_settings", methods=["GET"])
def credentials_get_settings():
    """
    Επιστρέφει τα stored general settings σε JSON — βολικό για AJAX αν το cog τα φορτώνει δυναμικά.
    """
    try:
        settings = load_settings() or {}
        return jsonify({"status": "ok", "settings": settings})
    except Exception as e:
        log.exception("Could not return settings")
        return jsonify({"status":"error","error":str(e)}), 500


@app.route("/api/check_mark", methods=["POST"])
def api_check_mark():
    """
    Payload JSON: { vat: str, mark: str }
    Επιστρέφει: { found: bool, source: "excel"/"none", characteristic: str|null, row: {...} }
    """
    try:
        payload = request.get_json(force=True)
        vat = payload.get("vat")
        mark = payload.get("mark")
        if not vat or not mark:
            return jsonify({"ok": False, "error": "missing vat or mark"}), 400

        safe_vat = secure_filename(vat)
        excel_path = os.path.join("uploads", f"{safe_vat}_invoices.xlsx")
        if not os.path.exists(excel_path):
            return jsonify({"ok": True, "found": False}), 200

        # διαβάζουμε το excel (single sheet)
        try:
            df = pd.read_excel(excel_path, dtype=str)  # read everything as str for safe compare
        except Exception as e:
            current_app.logger.exception("Failed to read excel for check_mark")
            return jsonify({"ok": False, "error": "cannot_read_excel", "detail": str(e)}), 500

        found, row, col = _match_row_by_mark(df, mark)
        if not found:
            return jsonify({"ok": True, "found": False}), 200

        # προσπαθούμε να βρούμε χαρακτηρισμό:
        # 1) πρώτα ψάχνουμε στο epsilon cache αν υπάρχει (καλύτερο για authoritative value)
        epsilon_path = os.path.join("data", "epsilon", f"{safe_vat}_epsilon_invoices.json")
        epsilon_list = _load_json(epsilon_path) or []
        # αναζητάμε στην epsilon λίστα για το ίδιο mark
        def _match_in_epsilon(item):
            for candidate_key in ("mark", "MARK", "invoice_id", "id", "Αριθμός Μητρώου", "Αριθμός"):
                if candidate_key in item and str(item[candidate_key]).strip() == str(mark).strip():
                    return True
            return False
        matched_eps = None
        for it in epsilon_list:
            if _match_in_epsilon(it):
                matched_eps = it
                break

        characteristic = None
        if matched_eps:
            # Έχουμε authoritative χαρακτηρισμό στον epsilon cache (προτεραιότητα)
            characteristic = matched_eps.get("χαρακτηρισμός") or matched_eps.get("characteristic") or matched_eps.get("flag")
        else:
            # αλλιώς ελέγχουμε αν υπάρχει στήλη χαρακτηρισμός στο excel row
            for candidate in ("χαρακτηρισμός", "χαρακτηρισμος", "characteristic", "flag", "Χαρακτηρισμός"):
                if candidate in row and row[candidate] not in (None, "", "nan"):
                    characteristic = row[candidate]
                    break

        return jsonify({
            "ok": True,
            "found": True,
            "source": "excel",
            "excel_column_used": col,
            "row": row,
            "characteristic": characteristic
        }), 200

    except Exception as e:
        current_app.logger.exception("api_check_mark failed")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/update_epsilon_characteristic", methods=["POST"])
def api_update_epsilon_characteristic():
    """
    Payload JSON: { vat: str, mark: str, characteristic: str }
    Ενημερώνει ΜΟΝΟ το data/epsilon/{vat}_epsilon_invoices.json το πεδίο χαρακτηρισμός
    (δημιουργεί εγγραφή αν δεν υπάρχει).
    """
    try:
        payload = request.get_json(force=True)
        vat = payload.get("vat")
        mark = payload.get("mark")
        new_char = payload.get("characteristic")
        if not vat or not mark:
            return jsonify({"ok": False, "error": "missing vat or mark"}), 400

        safe_vat = secure_filename(vat)
        epsilon_dir = os.path.join("data", "epsilon")
        os.makedirs(epsilon_dir, exist_ok=True)
        epsilon_path = os.path.join(epsilon_dir, f"{safe_vat}_epsilon_invoices.json")
        epsilon_list = _load_json(epsilon_path) or []

        # Try to find existing invoice by common keys
        def _match(item):
            for candidate_key in ("mark", "MARK", "invoice_id", "id", "Αριθμός Μητρώου", "Αριθμός"):
                if candidate_key in item and str(item[candidate_key]).strip() == str(mark).strip():
                    return True
            return False

        found = False
        for i, item in enumerate(epsilon_list):
            if _match(item):
                # update only χαρακτηρισμός-related keys
                epsilon_list[i]["χαρακτηρισμός"] = new_char
                epsilon_list[i]["characteristic"] = new_char  # για συμβατότητα
                epsilon_list[i]["_updated_at"] = datetime.utcnow().isoformat() + "Z"
                found = True
                break

        if not found:
            # Δημιουργούμε ελάχιστη εγγραφή μέσα στην epsilon cache (χωρίς άγγιγμα Excel)
            new_item = {
                "mark": mark,
                "χαρακτηρισμός": new_char,
                "characteristic": new_char,
                "_created_at": datetime.utcnow().isoformat() + "Z"
            }
            epsilon_list.append(new_item)

        _save_json(epsilon_path, epsilon_list)
        return jsonify({"ok": True, "updated": True, "found_existing": found}), 200

    except Exception as e:
        current_app.logger.exception("api_update_epsilon_characteristic failed")
        return jsonify({"ok": False, "error": str(e)}), 500


# ---------------- MARK search ----------------
@app.route("/search", methods=["GET", "POST"])
def search():
    result = None
    error = None
    mark = ""
    modal_summary = None
    invoice_lines = []
    customer_categories = []
    allow_edit_existing = False
    table_html = ""
    file_exists = False
    css_numcols = ""
    modal_warning = None
    fiscal_mismatch_block = False

    # ---------- helpers ----------
    def _map_invoice_type_local(code):
        INVOICE_TYPE_MAP = {
            "1.1": "Τιμολόγιο Πώλησης",
            "1.2": "Τιμολόγιο Πώλησης / Ενδοκοινοτικές Παραδόσεις",
            "1.3": "Τιμολόγιο Πώλησης / Παραδόσεις Τρίτων Χωρών",
            "1.4": "Τιμολόγιο Πώλησης / Πώληση για Λογαριασμό Τρίτων",
            "1.5": "Τιμολόγιο Πώλησης / Εκκαθάριση Πωλήσεων Τρίτων - Αμοιβή από Πωλήσεις Τρίτων",
            "1.6": "Τιμολόγιο Πώλησης / Συμπληρωματικό Παραστατικό",
            "2.1": "Τιμολόγιο Παροχής Υπηρεσιών",
            "2.2": "Τιμολόγιο Παροχής / Ενδοκοινοτική Παροχή Υπηρεσιών",
            "2.3": "Τιμολόγιο Παροχής / Παροχή Υπηρεσιών σε λήπτη Τρίτης Χώρας",
            "2.4": "Τιμολόγιο Παροχής / Συμπληρωματικό Παραστατικό",
            "3.1": "Τίτλος Κτήσης (μη υπόχρεος Εκδότης)",
            "3.2": "Τίτλος Κτήσης (άρνηση έκδοσης από υπόχρεο Εκδότη)",
            "5.1": "Πιστωτικό Τιμολόγιο / Συσχετιζόμενο",
            "5.2": "Πιστωτικό Τιμολόγιο / Μη Συσχετιζόμενο",
            "6.1": "Στοιχείο Αυτοπαράδοσης",
            "6.2": "Στοιχείο Ιδιοχρησιμοποίησης",
            "7.1": "Συμβόλαιο - Έσοδο",
            "8.1": "Ενοίκια - Έσοδο",
            "8.2": "Τέλος ανθεκτικότητας κλιματικής κρίσης",
            "8.4": "Απόδειξη Είσπραξης POS",
            "8.5": "Απόδειξη Επιστροφής POS",
            "8.6": "Δελτίο Παραγγελίας Εστίασης",
            "9.3": "Δελτίο Αποστολής",
            "11.1": "ΑΛΠ",
            "11.2": "ΑΠΥ",
            "11.3": "Απλοποιημένο Τιμολόγιο",
            "11.4": "Πιστωτικό Στοιχείο Λιανικής",
            "11.5": "Απόδειξη Λιανικής Πώλησης για Λογαριασμό Τρίτων",
            "13.1": "Έξοδα - Αγορές Λιανικών Συναλλαγών ημεδαπής / αλλοδαπής",
            "13.2": "Παροχή Λιανικών Συναλλαγών ημεδαπής / αλλοδαπής",
            "13.3": "Κοινόχρηστα",
            "13.4": "Συνδρομές",
            "13.30": "Παραστατικά Οντότητας ως Αναγράφονται από την ίδια (Δυναμικό)",
            "13.31": "Πιστωτικό Στοιχείο Λιανικής ημεδαπής / αλλοδαπής",
            "14.1": "Τιμολόγιο / Ενδοκοινοτικές Αποκτήσεις",
            "14.2": "Τιμολόγιο / Αποκτήσεις Τρίτων Χωρών",
            "14.3": "Τιμολόγιο / Ενδοκοινοτική Λήψη Υπηρεσιών",
            "14.4": "Τιμολόγιο / Λήψη Υπηρεσιών Τρίτων Χωρών",
            "14.5": "ΕΦΚΑ και λοιποί Ασφαλιστικοί Οργανισμοι",
            "14.30": "Παραστατικά Οντότητας ως Αναγράφονται από την ίδια (Δυναμικό)",
            "14.31": "Πιστωτικό ημεδαπής / αλλοδαπής",
            "15.1": "Συμβόλαιο - Έξοδο",
            "16.1": "Ενοίκιο Έξοδο",
            "17.1": "Μισθοδοσία",
            "17.2": "Αποσβέσεις",
            "17.3": "Λοιπές Εγγραφές Τακτοποίησης Εσόδων - Λογιστική Βάση",
            "17.4": "Λοιπές Εγγραφές Τακτοποίησης Εσόδων - Φορολογική Βάση",
            "17.5": "Λοιπές Εγγραφές Τακτοποίησης Εξόδων - Λογιστική Βάση",
            "17.6": "Λοιπές Εγγραφές Τακτοποίησης Εξόδων - Φορολογική Βάση",
        }
        return INVOICE_TYPE_MAP.get(str(code), str(code) or "")

    mapper = globals().get("map_invoice_type", None) or _map_invoice_type_local

    def float_from_comma(value):
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    def pick(src: dict, *keys, default=""):
        for k in keys:
            if k in src and src.get(k) not in (None, ""):
                return src.get(k)
        return default

    # ---------- credentials helpers ----------
    def credentials_path():
        return os.path.join(DATA_DIR, "credentials.json")

    def read_credentials_list_local():
        try:
            p = credentials_path()
            return json_read(p) or []
        except Exception:
            log.exception("read_credentials_list_local failed")
            return []

    def write_credentials_list_local(creds_list):
        try:
            p = credentials_path()
            try:
                json_write(p, creds_list)
            except Exception:
                tmp = p + ".tmp"
                with open(tmp, "w", encoding="utf-8") as f:
                    json.dump(creds_list, f, ensure_ascii=False, indent=2)
                os.replace(tmp, p)
            create_excel_fn = globals().get("create_empty_excel_for_vat")
            for c in (creds_list or []):
                try:
                    if not isinstance(c, dict):
                        continue
                    vat_c = c.get("vat") or c.get("AFM") or c.get("tax_number")
                    if vat_c and create_excel_fn and callable(create_excel_fn):
                        try:
                            create_excel_fn(vat_c)
                        except Exception:
                            log.exception("write_credentials_list_local: create_empty_excel_for_vat failed for %s", vat_c)
                except Exception:
                    log.exception("write_credentials_list_local: error ensuring excel for one credential")
            return True
        except Exception:
            log.exception("write_credentials_list_local failed")
            return False

    def find_active_client_index_local(creds_list, vat_to_match=None):
        if not creds_list:
            return None
        if vat_to_match:
            for i, c in enumerate(creds_list):
                try:
                    if str(c.get("vat", "")).strip() == str(vat_to_match).strip():
                        return i
                except Exception:
                    continue
        for i, c in enumerate(creds_list):
            if c.get("active"):
                return i
        return 0

    # ---------- active credential ----------
    active_cred = get_active_credential_from_session()
    vat = active_cred.get("vat") if active_cred else None

    # load customer_categories default from active_cred (used for invoices)
    try:
        raw_tags = active_cred.get("expense_tags") if active_cred else None
        if isinstance(raw_tags, str):
            customer_categories = [t.strip() for t in raw_tags.split(",") if t.strip()]
        elif isinstance(raw_tags, list):
            customer_categories = raw_tags
        if not customer_categories:
            customer_categories = [
                "αγορες_εμπορευματων",
                "αγορες_α_υλων",
                "γενικες_δαπανες",
                "αμοιβες_τριτων",
                "δαπανες_χωρις_φπα"
            ]
    except Exception:
        customer_categories = [
            "αγορες_εμπορευματων",
            "αγορες_α_υλων",
            "γενικες_δαπανες",
            "αμοιβες_τριτων",
            "δαπανες_χωρις_φπα"
        ]

    # --- Handle JSON AJAX request to save repeat mapping ---
    if request.method == "POST" and request.is_json:
        data = request.get_json() or {}
        if data.get("action") == "save_repeat_entry":
            enabled = bool(data.get("enabled", False))
            mapping = data.get("mapping", {}) or {}
            creds = read_credentials_list_local()
            idx = find_active_client_index_local(creds, vat_to_match=vat)
            if idx is None:
                return jsonify({"ok": False, "error": "No credentials found to save."}), 400
            try:
                creds[idx]["repeat_entry"] = {"enabled": enabled, "mapping": mapping}
                ok = write_credentials_list_local(creds)
                if not ok:
                    return jsonify({"ok": False, "error": "Failed to write credentials file."}), 500
                return jsonify({"ok": True})
            except Exception as e:
                log.exception("Failed saving repeat entry mapping")
                return jsonify({"ok": False, "error": str(e)}), 500

    # emulate POST: GET ?mark=...&force_edit=1
    emulate_post = False
    if request.method == "GET" and request.args.get("mark") and request.args.get("force_edit"):
        mark = request.args.get("mark", "").strip()
        emulate_post = True

    if request.method == "POST" or emulate_post:
        if request.method == "POST" and not request.is_json:
            mark = request.form.get("mark", "").strip()

        import re
        from urllib.parse import urlparse
        # existing invoice scrapers
        from scraper import scrape_wedoconnect, scrape_mydatapi, scrape_einvoice, scrape_impact, scrape_epsilon
        # safe import of receipt scraper
        try:
            from scraper_receipt import detect_and_scrape as detect_and_scrape_receipt
        except Exception:
            detect_and_scrape_receipt = None

        input_is_url = re.match(r'^https?://', mark)
        if input_is_url:
            domain = urlparse(mark).netloc.lower()
            scraped_afm = None
            scraped_marks = []
            try:
                if "wedoconnect" in domain:
                    scraped_marks, scraped_afm = scrape_wedoconnect(mark)
                elif "mydatapi.aade.gr" in domain:
                    data = scrape_mydatapi(mark)
                    scraped_marks = [data.get("MARK", "N/A")]
                    scraped_afm = data.get("ΑΦΜ Πελάτη")
                elif "einvoice.s1ecos.gr" in domain:
                    scraped_marks, scraped_afm = scrape_einvoice(mark)
                elif "einvoice.impact.gr" in domain or "impact.gr" in domain:
                    scraped_marks = scrape_impact(mark)
                elif "epsilonnet.gr" in domain:
                    mark_val, scraped_afm, _ = scrape_epsilon(mark)
                    if mark_val:
                        scraped_marks = [mark_val]
                else:
                    # fallback try receipt detector
                    if detect_and_scrape_receipt:
                        try:
                            rd = detect_and_scrape_receipt(mark)
                            if isinstance(rd, dict) and rd.get("MARK"):
                                scraped_marks = [str(rd.get("MARK"))]
                                scraped_afm = rd.get("issuer_vat") or rd.get("issuer_afm")
                        except Exception:
                            log.exception("Receipt detect_and_scrape failed for URL %s", mark)
                    if not scraped_marks:
                        error = "Άγνωστο URL για scraping."
            except Exception as e:
                log.exception("Scraping failed for URL %s", mark)
                error = f"Αποτυχία ανάγνωσης URL: {str(e)}"

            if scraped_afm and vat and str(scraped_afm).strip() != str(vat).strip():
                modal_warning = f"Το URL επιστρέφει ΑΦΜ {scraped_afm}, διαφορετικό από τον ενεργό πελάτη {vat}."

            if scraped_marks:
                mark = scraped_marks[0]

        if not input_is_url:
            if not vat:
                error = "Επέλεξε πρώτα έναν πελάτη (ΑΦΜ) για αναζήτηση."
            elif not mark or not mark.isdigit() or len(mark) != 15:
                error = "Πρέπει να δώσεις έγκυρο 15ψήφιο MARK."

        if not error:
            # --- φορτώνουμε cache invoices ---
            customer_file = os.path.join(DATA_DIR, f"{vat}_invoices.json")
            try:
                cache = json_read(customer_file) or []
            except Exception:
                log.exception("Failed to read customer_file %s", customer_file)
                cache = []
            docs_for_mark = [d for d in cache if str(d.get("mark", "")).strip() == mark]

            # flag for already classified docs
            classified_flag = False
            classified_docs = [d for d in docs_for_mark if str(d.get("classification", "")).strip().lower() == "χαρακτηρισμενο"]
            if classified_docs:
                error = f"Το MARK {mark} είναι ήδη χαρακτηρισμένο στο invoices.json."
                classified_flag = True

            # check duplicate in excel
            try:
                excel_path = excel_path_for(vat=vat)
                if os.path.exists(excel_path):
                    import pandas as pd
                    df_check = pd.read_excel(excel_path, engine="openpyxl", dtype=str).fillna("")
                    if "MARK" in df_check.columns:
                        marks_in_excel = df_check["MARK"].astype(str).str.strip().tolist()
                        if mark in marks_in_excel:
                            allow_edit_existing = True
            except Exception:
                log.exception("Could not read Excel to check duplicate MARK")

            # If not in cache, try receipt scraper to produce a single doc
            if not docs_for_mark and detect_and_scrape_receipt:
                try:
                    rd = detect_and_scrape_receipt(mark)
                    if isinstance(rd, dict) and rd.get("MARK"):
                        docs_for_mark = [{
                            "mark": str(rd.get("MARK")),
                            "issueDate": rd.get("issue_date") or rd.get("issueDate") or "",
                            "totalValue": rd.get("total_amount") or rd.get("totalAmount") or rd.get("total_value") or "",
                            "AFM_issuer": rd.get("issuer_vat") or rd.get("issuer_afm") or "",
                            "Name_issuer": rd.get("issuer_name") or rd.get("issuer") or "",
                            "type": rd.get("doc_type") or "receipt",
                            "_scraper_source": rd.get("source") or "receipt_scraper",
                            "_is_receipt": True
                        }]
                        log.info("search: receipt scraper returned a doc for mark %s vat %s", mark, vat)
                except Exception:
                    log.exception("Receipt scraper failed for mark %s", mark)

            if not docs_for_mark:
                error = f"MARK {mark} όχι στην cache του πελάτη {vat}. Κάνε πρώτα Fetch."
            else:
                if not classified_flag:
                    try:
                        # fiscal year check
                        sel_year = None
                        try:
                            sel_year = get_active_fiscal_year()
                        except Exception:
                            sel_year = None

                        first = docs_for_mark[0]
                        issue_date_str = pick(first, "issueDate", "issue_date", "date", "issueDate") or ""
                        issue_year = None
                        if issue_date_str and isinstance(issue_date_str, str):
                            m = re.search(r"\b(19|20)\d{2}\b", issue_date_str)
                            if m:
                                try:
                                    issue_year = int(m.group(0))
                                except Exception:
                                    issue_year = None
                        try:
                            sel_year_int = int(sel_year) if sel_year is not None else None
                        except Exception:
                            sel_year_int = None

                        if issue_year is not None and sel_year_int is not None and issue_year != sel_year_int:
                            modal_warning = (
                                f"Προσοχή: Το παραστατικό φαίνεται ότι εκδόθηκε το έτος {issue_year}, "
                                f"ενώ έχεις επιλεγμένη χρήση {sel_year_int}. Η διαδικασία μπλοκάρεται — "
                                "έλεγξε την ημερομηνία ή άλλαξε επιλεγμένη χρήση πριν συνεχίσεις."
                            )
                            fiscal_mismatch_block = True
                            log.info("search: fiscal year mismatch for MARK %s vat %s invoice_year=%s selected_year=%s", mark, vat, issue_year, sel_year_int)
                            modal_summary = None
                            invoice_lines = []
                            # keep customer_categories as default but don't allow edit
                            allow_edit_existing = False
                        else:
                            # detect receipts vs invoices
                            RECEIPT_TYPE_CODES = {"8.4", "8.5", "11.5"}
                            def is_receipt_doc(d):
                                try:
                                    if d.get("_is_receipt"):
                                        return True
                                    t = str(pick(d, "type", "invoiceType", "docType", default="")).strip()
                                    if t in RECEIPT_TYPE_CODES or t.lower() in ("receipt", "apodeixis", "απόδειξη"):
                                        return True
                                    desc = str(pick(d, "description", "desc", "Name", "Name_issuer", default="")).lower()
                                    if "απόδειξη" in desc or "receipt" in desc or "λιανική" in desc or "pos" in desc:
                                        return True
                                    has_net = any(k in d for k in ("totalNetValue", "totalNet", "lineTotal", "net"))
                                    has_vat = any(k in d for k in ("totalVatAmount", "totalVat", "vat", "vatRate"))
                                    if (not has_net) and (("totalValue" in d) or ("amount" in d)) and (not has_vat):
                                        return True
                                except Exception:
                                    pass
                                return False

                            receipt_votes = sum(1 for dd in docs_for_mark if is_receipt_doc(dd))
                            is_group_receipt = (receipt_votes >= max(1, len(docs_for_mark) // 2))

                            if is_group_receipt:
                                # Build receipt modal
                                invoice_lines = []
                                for idx, inst in enumerate(docs_for_mark):
                                    line_id = inst.get("id") or inst.get("line_id") or inst.get("LineId") or f"{mark}_inst{idx}"
                                    amount_raw = pick(inst, "totalValue", "amount", "lineTotal", "total", default=pick(inst, "value", "price", default=0))
                                    vat_raw = pick(inst, "totalVatAmount", "vat", "vatRate", default="")
                                    desc = pick(inst, "description", "desc", "Name", "note", default=f"Απόδειξη #{idx+1}")
                                    raw_vatcat = pick(inst, "vatCategory", "vat_category", "vatClass", "VATCategory", default="")
                                    mapped_vatcat = VAT_MAP.get(str(raw_vatcat).strip(), raw_vatcat) if raw_vatcat else ""
                                    invoice_lines.append({
                                        "id": line_id,
                                        "description": desc,
                                        "amount": amount_raw,
                                        "vat": vat_raw,
                                        "category": "",
                                        "vatCategory": mapped_vatcat
                                    })

                                total_net = 0.0
                                total_vat = 0.0
                                total_value = 0.0
                                for ln in invoice_lines:
                                    v = float_from_comma(ln.get("amount", 0))
                                    total_value += v
                                    vv = float_from_comma(ln.get("vat", 0))
                                    total_vat += vv
                                total_net = total_value - total_vat if total_vat else total_value

                                def fmt(x):
                                    try:
                                        return f"{x:.2f}".replace(".", ",")
                                    except Exception:
                                        return str(x)

                                for ml in invoice_lines:
                                    try:
                                        ml_v = float_from_comma(ml.get("amount", 0))
                                        ml["amount"] = fmt(ml_v)
                                    except Exception:
                                        pass
                                    try:
                                        mv = float_from_comma(ml.get("vat", 0))
                                        ml["vat"] = fmt(mv)
                                    except Exception:
                                        pass

                                modal_summary = {
                                    "mark": mark,
                                    "AA": pick(docs_for_mark[0], "AA", "aa", default=""),
                                    "AFM": pick(docs_for_mark[0], "AFM_issuer", "AFM", default=vat),
                                    "Name": pick(docs_for_mark[0], "Name_issuer", "Name", default=""),
                                    "series": pick(docs_for_mark[0], "series", "Series", default=""),
                                    "number": pick(docs_for_mark[0], "number", "aa", "AA", default=""),
                                    "issueDate": pick(docs_for_mark[0], "issueDate", "issue_date", default=""),
                                    "totalNetValue": fmt(total_net),
                                    "totalVatAmount": fmt(total_vat),
                                    "totalValue": fmt(total_value),
                                    "type": "receipt",
                                    "type_name": "Απόδειξη",
                                    "lines": invoice_lines,
                                    "is_receipt": True,
                                    # lock classification for receipts
                                    "χαρακτηρισμός": "αποδειξακια"
                                }

                                # receipts should NOT show repeat-entry editing / categories
                                customer_categories = []

                                # prefill from epsilon if exists (merge per-line categories if any)
                                try:
                                    epsilon_path = os.path.join(DATA_DIR, "epsilon", f"{vat}_epsilon_invoices.json")
                                    eps_list = json_read(epsilon_path) or []
                                    matched = None
                                    for it in (eps_list or []):
                                        try:
                                            if str(it.get("mark", "")).strip() == str(mark).strip():
                                                matched = it
                                                break
                                        except Exception:
                                            continue
                                    if matched:
                                        modal_summary['χαρακτηρισμός'] = matched.get('χαρακτηρισμός') or matched.get('characteristic') or modal_summary.get('χαρακτηρισμός', '') or "αποδειξακια"
                                        eps_lines = matched.get("lines", []) or []
                                        eps_by_id = {str(l.get("id", "")): l for l in eps_lines if l.get("id") is not None}
                                        for ml in modal_summary.get("lines", []):
                                            lid = str(ml.get("id", ""))
                                            if not lid:
                                                continue
                                            eps_line = eps_by_id.get(lid)
                                            if eps_line:
                                                if not ml.get("category") and eps_line.get("category"):
                                                    ml["category"] = eps_line.get("category")
                                                if (not ml.get("vatCategory") or ml.get("vatCategory") == "") and eps_line.get("vat_category"):
                                                    ml["vatCategory"] = eps_line.get("vat_category")
                                except Exception:
                                    log.exception("Could not prefill epsilon cache for receipts")

                            else:
                                # Invoice flow (restore categories selection)
                                invoice_lines = []
                                for idx, inst in enumerate(docs_for_mark):
                                    line_id = inst.get("id") or inst.get("line_id") or inst.get("LineId") or f"{mark}_inst{idx}"
                                    description = pick(inst, "description", "desc", "Description", "Name", "Name_issuer") or f"Instance #{idx+1}"
                                    amount = pick(inst, "amount", "lineTotal", "totalNetValue", "totalValue", "value", default="")
                                    vat_rate = pick(inst, "vat", "vatRate", "vatPercent", "totalVatAmount", default="")
                                    raw_vatcat = pick(inst, "vatCategory", "vat_category", "vatClass", "vatCategoryCode", "VATCategory", "vatCat", default="")
                                    mapped_vatcat = VAT_MAP.get(str(raw_vatcat).strip(), raw_vatcat) if raw_vatcat else ""
                                    invoice_lines.append({
                                        "id": line_id,
                                        "description": description,
                                        "amount": amount,
                                        "vat": vat_rate,
                                        "category": "",
                                        "vatCategory": mapped_vatcat
                                    })

                                first = docs_for_mark[0]
                                total_net = sum(float_from_comma(pick(d, "totalNetValue", "totalNet", "lineTotal", default=0)) for d in docs_for_mark)
                                total_vat = sum(float_from_comma(pick(d, "totalVatAmount", "totalVat", default=0)) for d in docs_for_mark)
                                total_value = total_net + total_vat

                                NEGATIVE_TYPES = {"5.1", "5.2", "11.4"}
                                inv_type = str(pick(first, "type", "invoiceType", default="")).strip()
                                is_negative = inv_type in NEGATIVE_TYPES

                                for ml in invoice_lines:
                                    try:
                                        v = float_from_comma(ml.get("amount", 0))
                                        ml["amount"] = f"{-abs(v):.2f}".replace(".", ",") if is_negative else f"{v:.2f}".replace(".", ",")
                                    except Exception:
                                        pass
                                    try:
                                        vv = float_from_comma(ml.get("vat", 0))
                                        ml["vat"] = f"{-abs(vv):.2f}".replace(".", ",") if is_negative else f"{vv:.2f}".replace(".", ",")
                                    except Exception:
                                        pass

                                modal_summary = {
                                    "mark": mark,
                                    "AA": pick(first, "AA", "aa", default=""),
                                    "AFM": pick(first, "AFM_issuer", "AFM_issuer", default=vat),
                                    "Name": pick(first, "Name", "Name_issuer", default=""),
                                    "series": pick(first, "series", "Series", "serie", default=""),
                                    "number": pick(first, "number", "aa", "AA", default=""),
                                    "issueDate": pick(first, "issueDate", "issue_date", default=pick(first, "issueDate", "issue_date", "")),
                                    "totalNetValue": (f"-{abs(total_net):.2f}" if is_negative else f"{total_net:.2f}").replace(".", ","),
                                    "totalVatAmount": (f"-{abs(total_vat):.2f}" if is_negative else f"{total_vat:.2f}").replace(".", ","),
                                    "totalValue": (f"-{abs(total_value):.2f}" if is_negative else f"{total_value:.2f}").replace(".", ","),
                                    "type": inv_type,
                                    "type_name": mapper(inv_type),
                                    "lines": invoice_lines,
                                    "is_receipt": False
                                }

                                # preserve existing epsilon prefill behavior for invoices
                                try:
                                    epsilon_path = os.path.join(DATA_DIR, "epsilon", f"{vat}_epsilon_invoices.json")
                                    eps_list = json_read(epsilon_path) or []

                                    if not eps_list:
                                        try:
                                            has_detail = False
                                            if modal_summary and (modal_summary.get("issueDate") or modal_summary.get("AFM") or modal_summary.get("AFM_issuer")):
                                                has_detail = True
                                            if invoice_lines and any((ln.get("description") or ln.get("amount") or ln.get("vat")) for ln in invoice_lines):
                                                has_detail = True
                                            if has_detail and modal_summary:
                                                epsilon_entry = {
                                                    "mark": str(modal_summary.get("mark", "")).strip(),
                                                    "issueDate": modal_summary.get("issueDate", "") or "",
                                                    "series": modal_summary.get("series", "") or "",
                                                    "aa": modal_summary.get("number", "") or modal_summary.get("AA", ""),
                                                    "AA": modal_summary.get("number", "") or modal_summary.get("AA", ""),
                                                    "type": modal_summary.get("type", "") or "",
                                                    "vatCategory": modal_summary.get("vatCategory", "") or "",
                                                    "totalNetValue": modal_summary.get("totalNetValue", "") or "",
                                                    "totalVatAmount": modal_summary.get("totalVatAmount", "") or "",
                                                    "totalValue": modal_summary.get("totalValue", "") or "",
                                                    "classification": modal_summary.get("classification", "") or "",
                                                    "χαρακτηρισμός": modal_summary.get("χαρακτηρισμός") or modal_summary.get("characteristic") or "",
                                                    "characteristic": modal_summary.get("χαρακτηρισμός") or modal_summary.get("characteristic") or "",
                                                    "AFM_issuer": modal_summary.get("AFM_issuer", "") or modal_summary.get("AFM", "") or "",
                                                    "Name_issuer": modal_summary.get("Name", "") or modal_summary.get("Name_issuer", "") or "",
                                                    "AFM": modal_summary.get("AFM", "") or vat,
                                                    "lines": []
                                                }
                                                for ln in invoice_lines:
                                                    epsilon_entry["lines"].append({
                                                        "id": ln.get("id", ""),
                                                        "description": ln.get("description", ""),
                                                        "amount": ln.get("amount", ""),
                                                        "vat": ln.get("vat", ""),
                                                        "category": ln.get("category", "") or "",
                                                        "vat_category": ln.get("vatCategory", "") or ""
                                                    })
                                                eps_list.append(epsilon_entry)
                                                try:
                                                    _safe_save_epsilon_cache(vat, eps_list)
                                                    log.info("Built per-mark epsilon entry for mark %s (vat %s)", epsilon_entry.get("mark"), vat)
                                                except Exception:
                                                    log.exception("Failed to save newly built per-mark epsilon entry")
                                        except Exception:
                                            log.exception("Failed building per-mark epsilon entry")

                                    # prefill from matched eps entry
                                    def match_epsilon_item(item, mark_val):
                                        for k in ('mark', 'MARK', 'invoice_id', 'Αριθμός Μητρώου', 'id'):
                                            if k in item and str(item[k]).strip() == str(mark_val).strip():
                                                return True
                                        return False

                                    matched = None
                                    for it in (eps_list or []):
                                        try:
                                            if match_epsilon_item(it, mark):
                                                matched = it
                                                break
                                        except Exception:
                                            continue

                                    if matched:
                                        modal_summary['χαρακτηρισμός'] = matched.get('χαρακτηρισμός') or matched.get('characteristic') or modal_summary.get('χαρακτηρισμός', '') or ""
                                        try:
                                            eps_lines = matched.get("lines", []) or []
                                            eps_by_id = {str(l.get("id", "")): l for l in eps_lines if l.get("id") is not None}
                                            for ml in modal_summary.get("lines", []):
                                                lid = str(ml.get("id", ""))
                                                if not lid:
                                                    continue
                                                eps_line = eps_by_id.get(lid)
                                                if eps_line:
                                                    if not ml.get("category") and eps_line.get("category"):
                                                        ml["category"] = eps_line.get("category")
                                                    if (not ml.get("vatCategory") or ml.get("vatCategory") == "") and eps_line.get("vat_category"):
                                                        ml["vatCategory"] = eps_line.get("vat_category")
                                        except Exception:
                                            log.exception("Failed to merge per-line categories from epsilon")
                                except Exception:
                                    log.exception("Could not read epsilon cache for prefill")

                    except Exception:
                        log.exception("search: fiscal year validation or modal build failed")
                        modal_summary = None
                        invoice_lines = []
                        # restore categories from defaults already loaded above
                        try:
                            raw_tags = active_cred.get("expense_tags") if active_cred else None
                            if isinstance(raw_tags, str):
                                customer_categories = [t.strip() for t in raw_tags.split(",") if t.strip()]
                            elif isinstance(raw_tags, list):
                                customer_categories = raw_tags
                            if not customer_categories:
                                customer_categories = [
                                    "αγορες_εμπορευματων",
                                    "αγορες_α_υλων",
                                    "γενικες_δαπανες",
                                    "αμοιβες_τριτων",
                                    "δαπανες_χωρις_φπα"
                                ]
                        except Exception:
                            customer_categories = [
                                "αγορες_εμπορευματων",
                                "αγορες_α_υλων",
                                "γενικες_δαπανες",
                                "αμοιβες_τριτων",
                                "δαπανες_χωρις_φπα"
                            ]
                else:
                    modal_summary = None
                    invoice_lines = []
                    # categories remain default
    # ---------- Read credentials to extract repeat_entry config to pass to template ----------
    repeat_entry_conf = {}
    try:
        creds_list = read_credentials_list_local()
        idx = find_active_client_index_local(creds_list, vat_to_match=vat)
        if idx is not None and idx < len(creds_list):
            repeat_entry_conf = creds_list[idx].get("repeat_entry", {}) or {}
    except Exception:
        log.exception("Could not read repeat_entry config from credentials")

    # --- Build table_html same way as /list ---
    try:
        active = get_active_credential_from_session()
        excel_path = DEFAULT_EXCEL_FILE
        if active and active.get("vat"):
            excel_path = excel_path_for(vat=active.get("vat"))
        elif active and active.get("name"):
            excel_path = excel_path_for(cred_name=active.get("name"))
        if os.path.exists(excel_path):
            file_exists = True
            import pandas as pd
            df = pd.read_excel(excel_path, engine="openpyxl", dtype=str).fillna("")
            df = df.astype(str)
            if "ΦΠΑ_ΑΝΑΛΥΣΗ" in df.columns:
                df = df.drop(columns=["ΦΠΑ_ΑΝΑΛΥΣΗ"])
            checkbox_html = '<input type="checkbox" name="delete_mark" />'
            df.insert(0, "✓", [checkbox_html] * len(df))
            table_html = df.to_html(classes="summary-table", index=False, escape=False)
        else:
            file_exists = False
            table_html = ""
    except Exception:
        log.exception("Failed building table_html for search page")
        table_html = ""

    return safe_render(
        "search.html",
        result=result,
        error=error,
        mark=mark,
        modal_summary=modal_summary,
        invoice_lines=invoice_lines,
        customer_categories=customer_categories,
        allow_edit_existing=allow_edit_existing,
        vat=vat,
        active_page="search",
        table_html=table_html,
        file_exists=file_exists,
        css_numcols=css_numcols,
        modal_warning=modal_warning,
        fiscal_mismatch_block=fiscal_mismatch_block,
        repeat_entry_conf=repeat_entry_conf
    )

@app.route("/api/next_receipt_mark", methods=["GET"])
def api_next_receipt_mark():
    """
    Επιστρέφει ένα 15ψήφιο επόμενο MARK για αποδείξεις.
    Βασίζεται στον ενεργό credential (get_active_credential_from_session),
    και ψάχνει υπάρχοντα MARKs σε:
      - excel_path_for(vat)
      - DATA_DIR/epsilon/<vat>_epsilon_invoices.json
      - DATA_DIR/<vat>_invoices.json
    Αν δεν βρεθεί τίποτα, ξεκινάει από DEFAULT_BASE_MARK (400000000000000).
    """
    try:
        # βοηθητικά
        def norm_mark_str(s):
            if s is None: return None
            s = str(s).strip()
            m = re.search(r"\d{15}", s)
            if m:
                return m.group(0)
            # try numeric conversion
            digits = re.sub(r"\D", "", s)
            if len(digits) >= 15:
                return digits[:15]
            if digits:
                return digits.zfill(15)
            return None

        # attempt to determine active VAT just like other endpoints
        vat = None
        try:
            active = get_active_credential_from_session()
            vat = active.get("vat") if active else None
        except Exception:
            vat = None

        existing_marks = set()

        # 1) read excel if exists
        try:
            if vat:
                path = excel_path_for(vat=vat)
                if path and os.path.exists(path):
                    import pandas as pd
                    df = pd.read_excel(path, engine="openpyxl", dtype=str).fillna("")
                    if "MARK" in df.columns:
                        for v in df["MARK"].astype(str).tolist():
                            s = norm_mark_str(v)
                            if s: existing_marks.add(s)
        except Exception:
            log.exception("api_next_receipt_mark: failed reading excel")

        # 2) read epsilon json for vat
        try:
            if vat:
                eps_path = os.path.join(DATA_DIR, "epsilon", f"{vat}_epsilon_invoices.json")
                if os.path.exists(eps_path):
                    try:
                        with open(eps_path, "r", encoding="utf-8") as f:
                            eps = json.load(f) or []
                        # eps might be list or dict
                        if isinstance(eps, dict):
                            # older shapes: keys maybe marks
                            for k in eps.keys():
                                s = norm_mark_str(k)
                                if s: existing_marks.add(s)
                            # also values may contain 'mark' fields
                            for v in eps.values():
                                try:
                                    if isinstance(v, dict):
                                        s = norm_mark_str(v.get("mark") or v.get("MARK") or v.get("invoice_id"))
                                        if s: existing_marks.add(s)
                                except Exception:
                                    pass
                        else:
                            for it in (eps or []):
                                try:
                                    if isinstance(it, dict):
                                        s = norm_mark_str(it.get("mark") or it.get("MARK") or it.get("invoice_id") or "")
                                        if s: existing_marks.add(s)
                                except Exception:
                                    pass
                    except Exception:
                        log.exception("api_next_receipt_mark: failed parsing epsilon json %s", eps_path)
        except Exception:
            log.exception("api_next_receipt_mark: epsilon read error")

        # 3) read cached invoices file
        try:
            if vat:
                cust_file = os.path.join(DATA_DIR, f"{vat}_invoices.json")
                if os.path.exists(cust_file):
                    try:
                        j = json_read(cust_file) or []
                        if isinstance(j, dict):
                            # if dict map by mark
                            for k in j.keys():
                                s = norm_mark_str(k)
                                if s: existing_marks.add(s)
                        else:
                            for it in (j or []):
                                try:
                                    if isinstance(it, dict):
                                        s = norm_mark_str(it.get("mark") or it.get("MARK") or "")
                                        if s: existing_marks.add(s)
                                except Exception:
                                    pass
                    except Exception:
                        log.exception("api_next_receipt_mark: failed parsing customer_file %s", cust_file)
        except Exception:
            log.exception("api_next_receipt_mark: cached invoices read error")

        # Convert to integers safely
        numeric_marks = []
        for m in existing_marks:
            try:
                numeric_marks.append(int(m))
            except Exception:
                continue

        DEFAULT_BASE_MARK = 400000000000000  # safe starting point if none found

        if numeric_marks:
            next_num = max(numeric_marks) + 1
        else:
            # try to seed from AFM or other heuristic — keep it deterministic
            try:
                if vat and vat.isdigit() and len(vat) <= 9:
                    # create a semi-deterministic seed: 400 + last digits of vat padded
                    seed_tail = vat.zfill(9)[-9:]
                    # simple mix to produce 15 digits: 400 + seed_tail + '00'
                    candidate = f"4{seed_tail}"  # might be <15
                    candidate = re.sub(r'\D', '', candidate).zfill(15)
                    next_num = int(candidate)
                    if next_num <= DEFAULT_BASE_MARK:
                        next_num = DEFAULT_BASE_MARK
                else:
                    next_num = DEFAULT_BASE_MARK
            except Exception:
                next_num = DEFAULT_BASE_MARK

        # ensure 15-digit string
        next_mark_str = str(next_num).zfill(15)
        if len(next_mark_str) > 15:
            next_mark_str = next_mark_str[-15:]

        return jsonify({"ok": True, "mark": next_mark_str})
    except Exception as e:
        log.exception("api_next_receipt_mark: unexpected")
        return jsonify({"ok": False, "error": str(e)}), 500

# --- Paste αυτό το endpoint μέσα στο app.py (δίπλα στα υπόλοιπα /api endpoints) ---
@app.route("/api/scrape_receipt", methods=["POST"])
def api_scrape_receipt():
    """
    Αναμένει JSON: { url: "<receipt url>" }
    Επιστρέφει: { ok: True, is_invoice: bool, mark: "<15digits or ''>", issue_date: "...", total_amount: "...", issuer_vat: "...", issuer_name: "...", progressive_aa: "...", raw: <original dict> }
    """
    try:
        data = request.get_json(silent=True) or {}
        url = data.get("url") or data.get("mark") or data.get("q") or ""
        if not url:
            return jsonify({"ok": False, "error": "Missing 'url' in request"}), 400

        log.info("api_scrape_receipt: incoming url=%s", url)

        # use scraper_receipt.detect_and_scrape (you said file scraper_receipt.py has it)
        try:
            from scraper_receipt import detect_and_scrape
        except Exception:
            # fallback to older function name if exists
            try:
                from scraper_receipt import detect_and_scrape as detect_and_scrape  # re-raise if not present
            except Exception as e:
                log.exception("api_scrape_receipt: cannot import detect_and_scrape")
                return jsonify({"ok": False, "error": "scraper_receipt detect_and_scrape not available"}), 500

        scraped = detect_and_scrape(url)
        # normalize expected shape (best-effort)
        if not scraped or not isinstance(scraped, dict):
            return jsonify({"ok": False, "error": "scraper returned unexpected result"}), 500

        # Example fields in your scraper_receipt output: MARK, doc_type, is_invoice, issue_date, issuer_name, issuer_vat, progressive_aa, total_amount, raw
        is_invoice = bool(scraped.get("is_invoice")) or False
        mark = str(scraped.get("MARK") or scraped.get("mark") or "").strip()
        issue_date = scraped.get("issue_date") or scraped.get("issueDate") or scraped.get("date") or ""
        total_amount = str(scraped.get("total_amount") or scraped.get("totalValue") or scraped.get("total") or scraped.get("totalValueGross") or "")
        issuer_vat = scraped.get("issuer_vat") or scraped.get("issuer_vat") or scraped.get("issuerAfm") or scraped.get("ΑΦΜ") or scraped.get("AFM") or ""
        issuer_name = scraped.get("issuer_name") or scraped.get("issuerName") or scraped.get("Name") or ""
        progressive_aa = scraped.get("progressive_aa") or scraped.get("AA") or scraped.get("aa") or ""

        log.info("api_scrape_receipt: scraped url=%s is_invoice=%s mark=%s", url, is_invoice, mark)

        return jsonify({
            "ok": True,
            "is_invoice": bool(is_invoice),
            "mark": mark,
            "issue_date": issue_date,
            "total_amount": total_amount,
            "issuer_vat": issuer_vat,
            "issuer_name": issuer_name,
            "progressive_aa": progressive_aa,
            "raw": scraped
        })
    except Exception as e:
        log.exception("api_scrape_receipt: unexpected")
        return jsonify({"ok": False, "error": str(e)}), 500




@app.route("/save_epsilon", methods=["POST"])
def save_epsilon():
    """
    Παρέχεται για απευθείας αποθήκευση epsilon (αν θέλεις ξεχωριστό κουμπί).
    Αναμένει form field "summary_json" (όπως το modal στέλνει) και αποθηκεύει
    το αντικείμενο στο per-vat epsilon file (data/epsilon/{vat}_epsilon_invoices.json).
    """
    active_cred = get_active_credential_from_session()
    if not active_cred:
        flash("Δεν υπάρχει ενεργός πελάτης για αποθήκευση.", "error")
        return redirect(url_for("search"))

    vat = active_cred.get("vat")
    if not vat:
        flash("Δεν βρέθηκε ΑΦΜ πελάτη.", "error")
        return redirect(url_for("search"))

    summary_json = request.form.get("summary_json")
    if not summary_json:
        flash("Δεν στάλθηκε δεδομένο για αποθήκευση.", "error")
        return redirect(url_for("search"))

    try:
        summary_data = json.loads(summary_json)
    except Exception as e:
        flash(f"Σφάλμα κατά την ανάγνωση του JSON: {e}", "error")
        return redirect(url_for("search"))

    # load existing epsilon cache and update per MARK
    epsilon_cache = load_epsilon_cache_for_vat(vat)

    mark = str(summary_data.get("mark", ""))
    existing_index = next((i for i, d in enumerate(epsilon_cache) if str(d.get("mark","")) == mark), None)
    if existing_index is not None:
        epsilon_cache[existing_index] = summary_data
    else:
        epsilon_cache.append(summary_data)

    # Save to disk
    try:
        save_epsilon_cache_for_vat(vat, epsilon_cache)
    except Exception:
        flash("Αποτυχία αποθήκευσης Epsilon αρχείου.", "error")
        return redirect(url_for("search"))

    flash(f"Το παραστατικό MARK {mark} αποθηκεύτηκε για Epsilon Excel.", "success")
    return redirect(url_for("search"))


@app.route("/save_accounts", methods=["POST"])
def save_accounts():
    """
    Αναμένει POST με πεδία:
      - account_<vat>__<expense_tag> = account_code
    Παράδειγμα πεδίου: account_24__γενικες_δαπανες = "70.02"
    Επιπλέον μπορεί να στέλνεται JSON payload.
    """
    try:
        # αν JSON payload
        if request.is_json:
            payload = request.get_json()
            accounts = payload.get("accounts", {})
            save_global_accounts_to_credentials(accounts)
            flash("Global accounts saved", "success")
            return redirect(url_for("credentials"))
        # αλλιώς form fields
        form = request.form
        # Δομή: accounts[vat][expense_tag] = code
        accounts = {}
        # αναζητούμε πεδία που ξεκινούν με "account_"
        for key in form:
            if not key.startswith("account_"):
                continue
            # key format: account_{vat}__{expense_tag}
            rest = key[len("account_"):]
            if "__" not in rest:
                continue
            vat_part, expense_tag = rest.split("__", 1)
            vat_key = vat_part.strip()
            code = form.get(key, "").strip()
            if vat_key not in accounts:
                accounts[vat_key] = {}
            accounts[vat_key][expense_tag] = code
        # αποθηκεύουμε
        save_global_accounts_to_credentials(accounts)
        flash("Global accounts saved", "success")
    except Exception as e:
        log.exception("save_accounts failed")
        flash(f"Could not save accounts: {e}", "error")
    return redirect(url_for("credentials"))

# ---------------- Save summary from modal to Excel & per-customer JSON ----------------
@app.route("/save_summary", methods=["POST"])
def save_summary():
    """
    Save summary and update epsilon:
      - αν υπάρχει ήδη detailed εγγραφή στο epsilon για το mark -> update-only per-line
      - αν υπάρχει μόνο placeholder για το mark -> διαγράφουμε τα placeholders και δημιουργούμε Excel + append
      - αλλιώς -> γράφει Excel και προσθέτει πλήρη εγγραφή στο epsilon

    Χρησιμοποιεί υπάρχοντες helpers του project: json_read/json_write, _safe_save_epsilon_cache,
    load_epsilon_cache_for_vat, append_summary_to_customer_file, excel_path_for, DEFAULT_EXCEL_FILE, VAT_MAP, κ.λπ.
    """
    try:
        from datetime import datetime as _dt, timezone as _tz
    except Exception:
        _dt = None
        _tz = None

    # canonical columns που θέλεις να εμφανίζονται στο Excel (προσαρμόζονται αν χρειαστεί)
    EXCEL_COLUMNS = [
        "MARK",
        "ΑΦΜ",
        "Επωνυμία",
        "Σειρά",
        "Αριθμός",
        "Ημερομηνία",
        "Είδος",
        "ΦΠΑ_ΚΑΤΗΓΟΡΙΑ",
        "Καθαρή Αξία",
        "ΦΠΑ",
        "Σύνολο"
    ]

    try:
        log.info("save_summary: start request from %s", request.remote_addr)
        # Προσπαθούμε πρώτα form field 'summary_json', μετά raw body (JSON), αλλιώς εναλλακτικά φτιάχνουμε minimal summary
        raw = None
        if request.form and request.form.get("summary_json"):
            raw = request.form.get("summary_json")
        else:
            # try JSON body
            try:
                # if content-type application/json, get_json will parse; else fallback to raw text
                if request.is_json:
                    summary = request.get_json(force=True)
                    raw = None  # parsed already
                else:
                    raw = request.get_data(as_text=True) or None
                    summary = None
            except Exception:
                raw = request.get_data(as_text=True) or None
                summary = None

        if raw:
            try:
                summary = json.loads(raw)
                log.debug("save_summary: parsed JSON summary keys: %s", list(summary.keys()))
            except Exception:
                log.warning("save_summary: failed json.loads(payload) - building minimal from form")
                summary = None

        if summary is None:
            # build minimal from form fields if possible
            summary = {}
            if request.form.get("mark"):
                summary["mark"] = request.form.get("mark")
                summary["AFM"] = request.form.get("vat") or request.form.get("AFM") or ""
            # optionally try other form keys
            if request.form.get("issueDate"):
                summary["issueDate"] = request.form.get("issueDate")
    except Exception:
        log.exception("save_summary: cannot parse payload")
        flash("Invalid summary payload", "error")
        return redirect(url_for("search"))

    # active VAT
    active = get_active_credential_from_session()
    vat = (active.get("vat") if active else None) or summary.get("AFM") or summary.get("AFM_issuer") or summary.get("AFM")
    if not vat:
        log.error("save_summary: missing vat - active=%s summary_afm=%s", bool(active), summary.get("AFM"))
        flash("No active customer selected (VAT)", "error")
        return redirect(url_for("search"))

    # helper to convert numeric strings like "1.234,56"
    def float_from_comma(value):
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip().replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0.0

    # --- ensure lines exist (reconstruct from per-customer invoices if needed) ---
    lines = summary.get("lines", []) or []
    if not lines:
        try:
            mark = str(summary.get("mark", "")).strip()
            docs_file = os.path.join(DATA_DIR, f"{vat}_invoices.json")
            all_docs = json_read(docs_file) or []
            docs_for_mark = [d for d in all_docs if str(d.get("mark", "")).strip() == mark]
            reconstructed = []
            for idx, inst in enumerate(docs_for_mark):
                ln_id = inst.get("id") or inst.get("line_id") or inst.get("LineId") or f"{mark}_inst{idx}"
                description = inst.get("description") or inst.get("desc") or inst.get("Description") or inst.get("Name") or inst.get("Name_issuer") or f"Instance #{idx+1}"
                amount = inst.get("amount") or inst.get("lineTotal") or inst.get("totalNetValue") or inst.get("totalValue") or ""
                vat_rate = inst.get("vat") or inst.get("vatRate") or inst.get("vatPercent") or inst.get("totalVatAmount") or ""
                raw_vatcat = inst.get("vatCategory") or inst.get("vat_category") or inst.get("VATCategory") or inst.get("vatCat") or ""
                if not raw_vatcat:
                    raw_vatcat = inst.get("vatCategoryCode") or inst.get("vat_code") or ""
                vat_cat_mapped = VAT_MAP.get(str(raw_vatcat).strip(), raw_vatcat) if raw_vatcat else ""
                reconstructed.append({
                    "id": ln_id,
                    "description": description,
                    "amount": amount,
                    "vat": vat_rate,
                    "category": "",
                    "vatCategory": vat_cat_mapped
                })
            lines = reconstructed
            log.debug("save_summary: reconstructed %d lines from %s", len(lines), docs_file)
        except Exception:
            log.exception("save_summary: failed to reconstruct lines")
            lines = []

    # --- normalize lines ---
    normalized_lines = []
    for idx, ln in enumerate(lines):
        if not isinstance(ln, dict):
            continue
        ln_id = ln.get("id") or f"{summary.get('mark','')}_l{idx}"
        raw_vcat = ln.get("vatCategory") or ln.get("vat_category") or ln.get("vatCat") or ln.get("vat_cat") or ""
        vcat_mapped = VAT_MAP.get(str(raw_vcat).strip(), raw_vcat) if raw_vcat else ""
        normalized_lines.append({
            "id": ln_id,
            "description": ln.get("description", "") or ln.get("desc", "") or "",
            "amount": ln.get("amount", "") or ln.get("lineTotal", "") or "",
            "vat": ln.get("vat", "") or ln.get("vatRate", "") or "",
            "category": ln.get("category", "") or "",
            "vatCategory": vcat_mapped
        })
    summary["lines"] = normalized_lines

    # --- append to per-customer summary JSON (best-effort) ---
    try:
        append_summary_to_customer_file(summary, vat)
    except Exception:
        log.exception("save_summary: append_summary_to_customer_file failed")

    # --- helper to detect placeholder vs detailed epsilon item ---
    def epsilon_item_has_detail(item):
        try:
            if not item or not isinstance(item, dict):
                return False
            if item.get("issueDate"):
                return True
            if item.get("totalNetValue") or item.get("totalValue"):
                return True
            lines_local = item.get("lines") or []
            if isinstance(lines_local, list):
                for l in lines_local:
                    if l and (l.get("description") or l.get("amount") or l.get("vat")):
                        return True
            if item.get("AFM_issuer") or item.get("AFM"):
                if item.get("aa") or item.get("AA") or item.get("issueDate") or item.get("totalValue"):
                    return True
        except Exception:
            pass
        return False

    # --- load epsilon cache for vat ---
    try:
        epsilon_cache = load_epsilon_cache_for_vat(vat)  # returns list or None
        if epsilon_cache is None:
            epsilon_cache = []
        log.debug("save_summary: loaded epsilon_cache len=%d for vat=%s", len(epsilon_cache), vat)
    except Exception:
        log.exception("save_summary: failed loading epsilon cache")
        epsilon_cache = []

    mark = str(summary.get("mark", "")).strip()

    # --- locate existing detailed entry + collect placeholders ---
    existing_index = None
    placeholder_indices = []
    try:
        for i, d in enumerate(epsilon_cache):
            try:
                d_mark = d.get("mark") or d.get("MARK") or d.get("invoice_id") or d.get("Αριθμός Μητρώου") or d.get("id")
                if d_mark is None or str(d_mark).strip() != mark:
                    continue
                if not epsilon_item_has_detail(d):
                    placeholder_indices.append(i)
                    log.debug("save_summary: found placeholder at idx=%d keys=%s", i, list(d.keys())[:6])
                    continue
                existing_index = i
                log.debug("save_summary: found detailed entry at idx=%d", i)
                break
            except Exception:
                log.exception("save_summary: error scanning epsilon_cache index %s", i)
                continue
    except Exception:
        log.exception("save_summary: scanning epsilon_cache failed unexpectedly")

    # --- if placeholders exist, remove them so we force Excel creation ---
    if placeholder_indices:
        try:
            for idx in sorted(placeholder_indices, reverse=True):
                try:
                    removed = epsilon_cache.pop(idx)
                    log.info("save_summary: removed placeholder epsilon entry idx=%d mark=%s vat=%s keys=%s", idx, mark, vat, list(removed.keys())[:6])
                except Exception:
                    log.exception("save_summary: failed to pop placeholder idx=%s", idx)
            try:
                _safe_save_epsilon_cache(vat, epsilon_cache)
                log.info("save_summary: persisted epsilon cache after removing placeholders for vat=%s", vat)
            except Exception:
                log.exception("save_summary: failed persisting epsilon cache after placeholder removal")
            existing_index = None
        except Exception:
            log.exception("save_summary: error removing placeholders")

    # --- if existing detailed entry -> update per-line categories only ---
    if existing_index is not None:
        try:
            existing = epsilon_cache[existing_index]
            existing_lines = existing.get("lines", []) or []
            existing_by_id = {str(l.get("id", "")): l for l in existing_lines if l.get("id") is not None}
            updated = False
            for ln in normalized_lines:
                lid = str(ln.get("id", ""))
                if not lid:
                    continue
                if lid in existing_by_id:
                    el = existing_by_id[lid]
                    new_cat = ln.get("category", "") or ""
                    if new_cat and str(el.get("category", "")) != new_cat:
                        el["category"] = new_cat
                        el["vat_category"] = ln.get("vatCategory", "") or el.get("vat_category", "")
                        updated = True
                    elif ln.get("vatCategory") and str(el.get("vat_category", "")) != ln.get("vatCategory"):
                        el["vat_category"] = ln.get("vatCategory")
                        updated = True
                else:
                    new_el = {
                        "id": lid,
                        "description": ln.get("description", ""),
                        "amount": ln.get("amount", ""),
                        "vat": ln.get("vat", ""),
                        "category": ln.get("category", "") or "",
                        "vat_category": ln.get("vatCategory", "") or ""
                    }
                    existing_lines.append(new_el)
                    existing_by_id[lid] = new_el
                    updated = True

            if updated:
                try:
                    if _dt and _tz:
                        existing["_updated_at"] = _dt.now(_tz.utc).isoformat()
                except Exception:
                    existing["_updated_at"] = existing.get("_updated_at", "")
                existing["lines"] = existing_lines
                epsilon_cache[existing_index] = existing

                try:
                    _safe_save_epsilon_cache(vat, epsilon_cache)
                    log.info("save_summary: updated epsilon per-line categories for mark=%s vat=%s", mark, vat)
                    flash("Ενημερώθηκε ο χαρακτηρισμός στο cache (epsilon).", "success")
                except Exception:
                    log.exception("save_summary: failed saving epsilon cache after update")
                    flash("Failed updating epsilon cache (see server logs)", "error")

                # ---- ensure Excel exists (if not, create it from summary/existing) ----
                try:
                    excel_path = excel_path_for(vat=vat)
                    if not os.path.exists(excel_path):
                        try:
                            total_net = float_from_comma(summary.get("totalNetValue", existing.get("totalNetValue", "") or 0))
                            total_vat = float_from_comma(summary.get("totalVatAmount", existing.get("totalVatAmount", "") or 0))
                        except Exception:
                            total_net = 0.0
                            total_vat = 0.0
                        total_value = total_net + total_vat
                        row = {
                            "MARK": str(summary.get("mark", existing.get("mark", ""))),
                            "ΑΦΜ": summary.get("AFM_issuer") or summary.get("AFM") or vat,
                            "Επωνυμία": summary.get("Name", existing.get("Name_issuer", "") or ""),
                            "Σειρά": summary.get("series", existing.get("series", "") or ""),
                            "Αριθμός": summary.get("number", existing.get("AA", existing.get("aa", ""))),
                            "Ημερομηνία": summary.get("issueDate", existing.get("issueDate", "")),
                            "Είδος": summary.get("type", existing.get("type", "")),
                            "ΦΠΑ_ΚΑΤΗΓΟΡΙΑ": summary.get("vatCategory", existing.get("vatCategory", "") or ""),
                            "Καθαρή Αξία": f"{total_net:.2f}".replace(".", ","),
                            "ΦΠΑ": f"{total_vat:.2f}".replace(".", ","),
                            "Σύνολο": f"{total_value:.2f}".replace(".", ",")
                        }
                        import pandas as pd
                        df_new = pd.DataFrame([row]).astype(str).fillna("")
                        # ensure canonical columns only
                        df_new = df_new.reindex(columns=EXCEL_COLUMNS, fill_value="")
                        os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
                        df_new.to_excel(excel_path, index=False, engine="openpyxl")
                        log.info("save_summary: existing epsilon updated but excel was missing -> created excel %s", excel_path)
                    else:
                        log.debug("save_summary: epsilon updated and excel already exists at %s; not re-writing", excel_path)
                except Exception:
                    log.exception("save_summary: failed ensuring/creating excel after epsilon update")

                return redirect(url_for("search"))
            else:
                log.info("save_summary: no per-line changes for mark=%s vat=%s", mark, vat)
                flash("Δεν υπήρξε αλλαγή στις κατηγορίες.", "info")
                return redirect(url_for("search"))
        except Exception:
            log.exception("save_summary: error processing existing detailed epsilon entry")
            flash("Server error while processing update", "error")
            return redirect(url_for("search"))

    # --- ELSE: no existing detailed entry -> create/write excel + append epsilon entry ---
    excel_path = excel_path_for(vat=vat)
    excel_written = False
    try:
        total_net = float_from_comma(summary.get("totalNetValue", 0))
        total_vat = float_from_comma(summary.get("totalVatAmount", 0))
        total_value = float_from_comma(summary.get("totalValue", total_net + total_vat))

        row = {
            "MARK": str(summary.get("mark", "")),
            "ΑΦΜ": summary.get("AFM_issuer") or summary.get("AFM") or vat,
            "Επωνυμία": summary.get("Name", ""),
            "Σειρά": summary.get("series", ""),
            "Αριθμός": summary.get("number", ""),
            "Ημερομηνία": summary.get("issueDate", ""),
            "Είδος": summary.get("type", ""),
            "ΦΠΑ_ΚΑΤΗΓΟΡΙΑ": summary.get("vatCategory", ""),
            "Καθαρή Αξία": f"{total_net:.2f}".replace(".", ","),
            "ΦΠΑ": f"{total_vat:.2f}".replace(".", ","),
            "Σύνολο": f"{total_value:.2f}".replace(".", ",")
        }

        import pandas as pd
        df_new = pd.DataFrame([row]).astype(str).fillna("")
        # ensure df_new has only canonical columns (and in correct order)
        df_new = df_new.reindex(columns=EXCEL_COLUMNS, fill_value="")

        # if there's a higher-level helper, prefer calling it (backward compat)
        helper = globals().get("_ensure_excel_and_update_or_append") or globals().get("_append_to_excel")
        if helper and callable(helper):
            try:
                log.debug("save_summary: calling helper %s for excel update", helper.__name__)
                # prefer passing summary if helper expects it, otherwise pass row+vat
                try:
                    helper(summary, vat=vat)
                except TypeError:
                    helper(row, vat=vat)
                excel_written = True
                log.info("save_summary: helper %s completed", helper.__name__)
            except Exception:
                log.exception("save_summary: helper %s failed; falling back to inline Excel write", helper.__name__)

        if not excel_written:
            if os.path.exists(excel_path):
                try:
                    df_existing = pd.read_excel(excel_path, engine="openpyxl", dtype=str).fillna("")
                    # coerce to strings and remove any non-canonical columns
                    df_existing = df_existing.astype(str).fillna("")
                    # keep only canonical columns and in canonical order (drops anything else)
                    df_existing = df_existing.reindex(columns=EXCEL_COLUMNS, fill_value="")
                    # concat and ensure final has canonical columns only
                    df_concat = pd.concat([df_existing, df_new], ignore_index=True, sort=False)
                    df_concat = df_concat.reindex(columns=EXCEL_COLUMNS, fill_value="")
                    df_concat.to_excel(excel_path, index=False, engine="openpyxl")
                    log.info("save_summary: appended row to existing excel %s", excel_path)
                except Exception:
                    log.exception("save_summary: failed to append to existing excel %s", excel_path)
                    raise
            else:
                try:
                    os.makedirs(os.path.dirname(excel_path) or ".", exist_ok=True)
                    df_new.to_excel(excel_path, index=False, engine="openpyxl")
                    log.info("save_summary: created new excel and wrote row %s", excel_path)
                except Exception:
                    log.exception("save_summary: failed to create/write new excel %s", excel_path)
                    raise
            flash("Saved to Excel.", "success")
    except Exception:
        log.exception("save_summary: Excel write failed")
        flash("Excel save failed", "error")

    # --- Build epsilon entry and append it to cache ---
    try:
        epsilon_entry = {
            "mark": mark,
            "issueDate": summary.get("issueDate", ""),
            "series": summary.get("series", ""),
            "aa": summary.get("number", "") or summary.get("AA", "") or summary.get("aa", ""),
            "AA": summary.get("number", "") or summary.get("AA", "") or summary.get("aa", ""),
            "type": summary.get("type", ""),
            "vatCategory": summary.get("vatCategory", ""),
            "totalNetValue": summary.get("totalNetValue", ""),
            "totalVatAmount": summary.get("totalVatAmount", ""),
            "totalValue": summary.get("totalValue", ""),
            "classification": summary.get("classification", ""),
            "χαρακτηρισμός": summary.get("χαρακτηρισμός") or summary.get("characteristic") or "",
            "characteristic": summary.get("χαρακτηρισμός") or summary.get("characteristic") or "",
            "AFM_issuer": summary.get("AFM_issuer", "") or summary.get("AFM", ""),
            "Name_issuer": summary.get("Name_issuer") or summary.get("Name", ""),
            "AFM": summary.get("AFM", "") or vat,
            "lines": []
        }
        for ln in normalized_lines:
            epsilon_entry["lines"].append({
                "id": ln.get("id", ""),
                "description": ln.get("description", ""),
                "amount": ln.get("amount", ""),
                "vat": ln.get("vat", ""),
                "category": ln.get("category", "") or "",
                "vat_category": ln.get("vatCategory", "") or ""
            })

        epsilon_cache.append(epsilon_entry)
        try:
            _safe_save_epsilon_cache(vat, epsilon_cache)
            log.info("save_summary: appended new epsilon entry for mark=%s vat=%s", mark, vat)
            flash("Saved summary and appended new epsilon entry.", "success")
        except Exception:
            log.exception("save_summary: failed saving new epsilon cache")
            flash("Failed updating epsilon cache (see server logs)", "error")
    except Exception:
        log.exception("save_summary: failed building/appending epsilon entry")
        flash("Failed updating epsilon cache", "error")

    return redirect(url_for("search"))


@app.route("/save_receipt", methods=["POST"])
def save_receipt():
    receipt = request.form.to_dict()
    vat = receipt.get("issuer_vat")
    year = receipt.get("issue_date")[-4:]
    category = receipt.get("category")

    # --- Αποθήκευση Excel ---
    excel_file = f"uploads/{vat}_{year}_invoices.xlsx"
    df_new = pd.DataFrame([receipt])
    if os.path.exists(excel_file):
        df_existing = pd.read_excel(excel_file)
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    else:
        df_combined = df_new
    df_combined.to_excel(excel_file, index=False)

    # --- Αποθήκευση JSON ---
    json_file = f"data/{vat}_epsilon_invoices.json"
    if os.path.exists(json_file):
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    else:
        data = []

    receipt["category"] = category  # αποθηκεύουμε τον server-side χαρακτηρισμό
    data.append(receipt)
    with open(json_file, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    flash("Η απόδειξη αποθηκεύτηκε επιτυχώς!", "success")
    return redirect(url_for("search"))



@app.route("/api/confirm_receipt", methods=["POST"])
def api_confirm_receipt():
    import os, json, traceback
    from datetime import datetime
    from filelock import FileLock
    import pandas as pd

    debug_dump_path = "/tmp/confirm_receipt_debug.json"
    try:
        payload = request.get_json(silent=True)
        if not payload:
            payload = request.form.to_dict() or {}

        # debug dump
        try:
            with open(debug_dump_path, "w", encoding="utf-8") as df:
                json.dump({"received_at": datetime.utcnow().isoformat()+"Z", "payload": payload}, df, ensure_ascii=False, indent=2)
        except Exception as e:
            app.logger.warning("Could not write debug dump: %s", e)

        app.logger.info("api_confirm_receipt: payload keys: %s", list(payload.keys()))

        summary = payload.get("summary") or {}
        # Extract important fields (robust)
        supplied_afm = (payload.get("afm") or payload.get("AFM") or "").strip()
        # try session active credential if not supplied
        client_afm = supplied_afm or ""
        if not client_afm:
            try:
                active = session.get("active_credential") if session else None
                if active and active.get("vat"):
                    client_afm = str(active.get("vat")).strip()
            except Exception:
                client_afm = client_afm

        if not client_afm:
            return jsonify({"ok": False, "error": "Missing AFM for active client (session)."}), 400

        # normalize summary values
        mark = (payload.get("mark") or summary.get("mark") or "").strip()
        issue_date = (summary.get("issueDate") or summary.get("issue_date") or "").strip()
        total_amount = (summary.get("totalValue") or summary.get("total_amount") or payload.get("total_amount") or "").strip()
        progressive_aa = (summary.get("AA") or summary.get("aa") or summary.get("progressive_aa") or "").strip()
        issuer_vat = (summary.get("AFM_issuer") or summary.get("AFM") or (summary.get("issuer_vat") if summary.get("issuer_vat") else "")).strip()
        issuer_name = (summary.get("Name") or summary.get("Name_issuer") or summary.get("issuer_name") or "").strip()
        doc_type = (summary.get("type") or summary.get("type_name") or summary.get("doc_type") or "").strip()
        # category we want receipts to have
        chosen_category = (payload.get("category") or summary.get("category") or "αποδειξακια").strip()

        # if the scraped data reports an invoice, refuse saving as receipt (frontend usually warns)
        if (summary.get("is_invoice") is True) or (str(doc_type).lower().startswith("1") and summary.get("is_invoice") is not True):
            return jsonify({"ok": False, "error": "Document is an invoice, not a receipt (will not save as receipt)."}), 400

        # derive year
        year = str(payload.get("year") or summary.get("year") or "")
        if not year:
            try:
                if issue_date:
                    parts = issue_date.split("/")
                    if len(parts) >= 3:
                        year = parts[-1]
                if not year:
                    year = str(datetime.utcnow().year)
            except Exception:
                year = str(datetime.utcnow().year)

        # ensure mark present (fallback to timestamp-based mark to avoid empty MARK)
        if not mark:
            mark = datetime.utcnow().strftime("%Y%m%d%H%M%S")  # fallback unique-ish mark

        # ensure directories
        base_data = os.path.join("data")
        excel_dir = os.path.join(base_data, "excel")
        epsilon_dir = os.path.join(base_data, "epsilon")
        os.makedirs(excel_dir, exist_ok=True)
        os.makedirs(epsilon_dir, exist_ok=True)

        # Use ONE excel per client AFM + fiscal year as you requested:
        excel_filename = f"{client_afm}_{year}_invoices.xlsx"
        excel_path = os.path.join(excel_dir, excel_filename)
        json_path = os.path.join(epsilon_dir, f"{client_afm}_epsilon_invoices.json")

        app.logger.info("api_confirm_receipt: saving for client_afm=%s year=%s mark=%s -> excel=%s json=%s",
                        client_afm, year, mark, excel_path, json_path)

        # --- Write Excel (append/replace row for MARK) ---
        try:
            lock_excel = FileLock(excel_path + ".lock")
            with lock_excel:
                if os.path.exists(excel_path):
                    try:
                        df = pd.read_excel(excel_path, engine="openpyxl", dtype=str).fillna("")
                    except Exception as e:
                        app.logger.warning("Failed reading existing excel, creating new. err=%s", e)
                        df = pd.DataFrame()
                else:
                    df = pd.DataFrame()

                # Desired columns for invoices/receipts sheet (keep compatibility)
                desired_cols = ["MARK", "issueDate", "AFM_issuer", "AA", "series", "totalNetValue", "totalVatAmount", "totalValue", "character", "type"]
                for c in desired_cols:
                    if c not in df.columns:
                        df[c] = ""

                # remove existing same MARK (robust)
                try:
                    df = df[~(df["MARK"].astype(str).fillna("").str.strip() == str(mark).strip())]
                except Exception:
                    df = df[df["MARK"].astype(str).fillna("") != str(mark)]

                row = {
                    "MARK": mark,
                    "issueDate": issue_date or "",
                    "AFM_issuer": issuer_vat or "",
                    "AA": progressive_aa or "",
                    "series": summary.get("series") or "",
                    # For receipts we often only have total -> store both net and total same
                    "totalNetValue": summary.get("totalNetValue") or total_amount or "",
                    "totalVatAmount": summary.get("totalVatAmount") or "",
                    "totalValue": summary.get("totalValue") or total_amount or "",
                    "character": chosen_category or "αποδειξακια",
                    "type": doc_type or (summary.get("type_name") or "")
                }
                # append keeping dataframe columns
                df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
                # write back (index=False keeps clean excel)
                df.to_excel(excel_path, index=False)
                app.logger.info("api_confirm_receipt: excel saved ok (%s rows)", len(df))
        except Exception as e:
            app.logger.exception("api_confirm_receipt: excel write failed: %s", e)
            # do not abort — attempt json write as well

        # --- Write JSON (list) ---
        try:
            lock_json = FileLock(json_path + ".lock")
            with lock_json:
                existing = []
                if os.path.exists(json_path):
                    try:
                        with open(json_path, "r", encoding="utf-8") as f:
                            existing = json.load(f)
                    except Exception as e:
                        app.logger.warning("Existing json parse failed, will reset to empty list. err=%s", e)
                        existing = []

                # normalize existing to list
                if isinstance(existing, dict):
                    converted = []
                    for k, v in existing.items():
                        if isinstance(v, dict):
                            converted.append(v)
                    existing = converted
                if not isinstance(existing, list):
                    existing = []

                # Build entry following example structure
                entry = {
                    "mark": mark,
                    "issueDate": issue_date or "",
                    "series": summary.get("series") or "",
                    "aa": progressive_aa or "",
                    "AA": progressive_aa or "",
                    "type": doc_type or summary.get("type_name") or "",
                    "vatCategory": summary.get("vatCategory") or "",
                    "totalNetValue": (summary.get("totalNetValue") or total_amount or ""),
                    "totalVatAmount": (summary.get("totalVatAmount") or ""),
                    "totalValue": (summary.get("totalValue") or total_amount or ""),
                    "classification": "",
                    "χαρακτηρισμός": chosen_category or "αποδειξακια",
                    "characteristic": chosen_category or "αποδειξακια",
                    "AFM_issuer": issuer_vat or "",
                    "Name_issuer": issuer_name or "",
                    "AFM": issuer_vat or "",
                    "Name": issuer_name or "",
                    "lines": []
                }

                # collect lines from summary (respect category if present; default 'αποδειξακια')
                incoming_lines = summary.get("lines") or summary.get("raw", {}).get("lines") or []
                if isinstance(incoming_lines, dict):
                    try:
                        incoming_lines = list(incoming_lines.values())
                    except Exception:
                        incoming_lines = []

                if not isinstance(incoming_lines, list) or len(incoming_lines) == 0:
                    # create single-line from total if no lines provided
                    entry["lines"].append({
                        "id": entry["mark"] + "_inst0",
                        "description": summary.get("Name") or summary.get("issuer_name") or "",
                        "amount": total_amount or "",
                        "vat": summary.get("totalVatAmount") or "",
                        "category": chosen_category or "αποδειξακια",
                        "vat_category": summary.get("vatCategory") or ""
                    })
                else:
                    for idx, ln in enumerate(incoming_lines):
                        try:
                            # accept both dict-like or objects
                            if not isinstance(ln, dict):
                                ln = dict(ln)
                        except Exception:
                            ln = {}
                        try:
                            lid = ln.get("id") or ln.get("line_id") or ln.get("LineId") or (f"{entry['mark']}_inst{idx}")
                            desc = ln.get("description") or ln.get("desc") or ln.get("descriptionText") or ""
                            amt = ln.get("amount") or ln.get("lineTotal") or ln.get("total") or ""
                            vat = ln.get("vat") or ln.get("vatRate") or ""
                            ln_cat = ln.get("category") or chosen_category or "αποδειξακια"
                            vat_cat = ln.get("vatCategory") or ln.get("vat_category") or ""
                            entry["lines"].append({
                                "id": lid,
                                "description": desc,
                                "amount": amt,
                                "vat": vat,
                                "category": ln_cat,
                                "vat_category": vat_cat
                            })
                        except Exception:
                            continue

                entry["_updated_at"] = datetime.utcnow().isoformat() + "+00:00"

                # Replace or append by mark
                replaced = False
                for i, it in enumerate(existing):
                    try:
                        if str((it.get("mark") or it.get("MARK") or "")).strip() == str(mark).strip():
                            existing[i] = entry
                            replaced = True
                            break
                    except Exception:
                        continue
                if not replaced:
                    existing.append(entry)

                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(existing, f, ensure_ascii=False, indent=2)
                app.logger.info("api_confirm_receipt: json saved ok (entries=%d)", len(existing))
        except Exception as e:
            app.logger.exception("api_confirm_receipt: json write failed: %s", e)
            return jsonify({"ok": False, "error": "json write failed: " + str(e)}), 500

        return jsonify({"ok": True, "excel": excel_path, "json": json_path})
    except Exception as e:
        app.logger.exception("api_confirm_receipt: unexpected error: %s", e)
        try:
            with open(debug_dump_path, "a", encoding="utf-8") as df:
                df.write("\n\nEXCEPTION:\n")
                df.write(traceback.format_exc())
        except Exception:
            pass
        return jsonify({"ok": False, "error": str(e)}), 500










# ---------------- List / download ----------------
@app.route("/list", methods=["GET"])
def list_invoices():
    # choose excel file based on active session credential
    active = get_active_credential_from_session()
    excel_path = DEFAULT_EXCEL_FILE
    if active and active.get("vat"):
        excel_path = excel_path_for(vat=active.get("vat"))
    elif active and active.get("name"):
        excel_path = excel_path_for(cred_name=active.get("name"))
    else:
        excel_path = DEFAULT_EXCEL_FILE

    if request.args.get("download") and os.path.exists(excel_path):
        # download the active client's file
        return send_file(excel_path, as_attachment=True, download_name=os.path.basename(excel_path),
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    table_html = ""
    error = ""
    css_numcols = ""

    if os.path.exists(excel_path):
        try:
            df = pd.read_excel(excel_path, engine="openpyxl", dtype=str).fillna("")
            df = df.astype(str)

            if "ΦΠΑ_ΑΝΑΛΥΣΗ" in df.columns:
                df = df.drop(columns=["ΦΠΑ_ΑΝΑΛΥΣΗ"])

            if "MARK" in df.columns:
                checkboxes = df["MARK"].apply(lambda v: f'<input type="checkbox" name="delete_mark" value="{str(v)}">')
                df.insert(0, "✓", checkboxes)

            table_html = df.to_html(classes="summary-table", index=False, escape=False)
            table_html = table_html.replace("<th>✓</th>", '<th><input type="checkbox" id="selectAll" title="Επιλογή όλων"></th>')
            table_html = table_html.replace("<td>", '<td><div class="cell-wrap">').replace("</td>", "</div></td>")

            import re
            headers = re.findall(r'<th[^>]*>(.*?)</th>', table_html, flags=re.S)
            num_indices = []
            for i, h in enumerate(headers):
                text = re.sub(r'<.*?>', '', h).strip()
                if text in ("Καθαρή Αξία", "ΦΠΑ", "Σύνολο", "Total", "Net", "VAT") or "ΦΠΑ" in text or "ΠΟΣΟ" in text:
                    num_indices.append(i+1)
            css_rules = []
            for idx in num_indices:
                css_rules.append(f".summary-table td:nth-child({idx}), .summary-table th:nth-child({idx}) {{ text-align: right; }}")
            css_numcols = "\n".join(css_rules)

        except Exception as e:
            error = f"Σφάλμα ανάγνωσης Excel: {e}"
    else:
        error = f"Δεν βρέθηκε το αρχείο {os.path.basename(excel_path)}."

    # pass active_credential name so navbar and templates can reflect it
    active_name = session.get("active_credential")
    return safe_render("list.html",
                       table_html=Markup(table_html),
                       error=error,
                       file_exists=os.path.exists(excel_path),
                       css_numcols=css_numcols,
                       active_page="list_invoices",
                       active_credential=active_name)

# ---------------- Delete invoices ----------------
@app.route("/delete", methods=["POST"])
def delete_invoices():
    """
    Delete selected MARKs:
      - removes rows from active customer's Excel file (if MARK column exists)
      - removes matching entries from per-VAT epsilon cache (epsilon/..._epsilon_invoices.json)
    DOES NOT modify the per-customer invoices.json file.
    Extensive logging for debugging.
    """
    try:
        log.info("delete_invoices: request from %s form_keys=%s", request.remote_addr, list(request.form.keys()))
    except Exception:
        pass

    # collect and normalize marks (primary)
    marks_to_delete = request.form.getlist("delete_mark") or []
    # fallback: maybe frontend sent JSON or comma-separated
    if not marks_to_delete:
        raw = request.form.get("delete_mark_json") or request.form.get("delete_marks") or request.form.get("marks")
        if raw:
            try:
                import json as _json
                parsed = _json.loads(raw)
                if isinstance(parsed, list):
                    marks_to_delete = parsed
            except Exception:
                marks_to_delete = [m.strip() for m in str(raw).split(",") if m.strip()]

    # normalize to strings and dedupe
    marks_to_delete = [str(m).strip() for m in marks_to_delete if str(m).strip()]
    marks_to_delete = list(dict.fromkeys(marks_to_delete))  # preserve order, dedupe

    log.info("delete_invoices: marks_to_delete resolved = %s", marks_to_delete)

    if not marks_to_delete:
        flash("Δεν επιλέχθηκε κανένα MARK για διαγραφή.", "error")
        return redirect(url_for("search"))

    # determine active customer's excel file
    active = get_active_credential_from_session()
    excel_path = DEFAULT_EXCEL_FILE
    if active and active.get("vat"):
        excel_path = excel_path_for(vat=active.get("vat"))
    elif active and active.get("name"):
        excel_path = excel_path_for(cred_name=active.get("name"))

    deleted_from_excel = 0
    try:
        if os.path.exists(excel_path):
            import pandas as pd
            df = pd.read_excel(excel_path, engine="openpyxl", dtype=str).fillna("")
            # normalize column names
            cols = [c.strip() for c in df.columns.astype(str)]
            df.columns = cols

            if "MARK" in df.columns:
                marks_series = df["MARK"].astype(str).str.strip()
                mask = marks_series.isin(marks_to_delete)
                num_matches = int(mask.sum())
                if num_matches > 0:
                    df_remaining = df[~mask].copy()
                    try:
                        # If no rows remain, write an empty dataframe (preserving columns)
                        if df_remaining.shape[0] == 0:
                            empty_df = df.iloc[0:0].copy()
                            empty_df.to_excel(excel_path, index=False, engine="openpyxl")
                        else:
                            df_remaining.to_excel(excel_path, index=False, engine="openpyxl")
                        deleted_from_excel = num_matches
                        log.info("delete_invoices: deleted %d marks from Excel %s: %s", num_matches, excel_path, marks_to_delete)
                    except Exception:
                        log.exception("delete_invoices: failed writing Excel after deletion %s", excel_path)
                else:
                    log.info("delete_invoices: no matching MARKs found in Excel %s for deletion: %s", excel_path, marks_to_delete)
            else:
                log.warning("delete_invoices: Excel file %s does not contain a 'MARK' column; skipping Excel deletion", excel_path)
        else:
            log.info("delete_invoices: Excel path %s does not exist; skipping Excel deletion.", excel_path)
    except Exception:
        log.exception("delete_invoices: Error while deleting from Excel")

    # delete matching entries from per-VAT epsilon cache ONLY
    deleted_from_epsilon = 0
    try:
        vat = active.get("vat") if active else None
        if vat:
            # try helper to get epsilon path, else guess
            try:
                epsilon_path = epsilon_file_path_for(vat)
            except Exception:
                epsilon_path = os.path.join(DATA_DIR, "epsilon", f"{vat}_epsilon_invoices.json")

            if os.path.exists(epsilon_path):
                try:
                    eps_cache = json_read(epsilon_path) or []
                except Exception:
                    try:
                        with open(epsilon_path, "r", encoding="utf-8") as f:
                            eps_cache = json.load(f) or []
                    except Exception:
                        eps_cache = []

                before_len = len(eps_cache)

                def item_mark_val(it):
                    for k in ("mark", "MARK", "invoice_id", "Αριθμός Μητρώου", "id"):
                        if isinstance(it, dict) and k in it and it.get(k) not in (None, ""):
                            return str(it.get(k)).strip()
                    return ""

                new_cache = [e for e in eps_cache if item_mark_val(e) not in marks_to_delete]
                after_len = len(new_cache)
                deleted_from_epsilon = before_len - after_len

                if deleted_from_epsilon > 0:
                    try:
                        # Prefer safe helper if available
                        if globals().get("_safe_save_epsilon_cache"):
                            _safe_save_epsilon_cache(vat, new_cache)
                        else:
                            try:
                                json_write(epsilon_path, new_cache)
                            except Exception:
                                tmp = epsilon_path + ".tmp"
                                with open(tmp, "w", encoding="utf-8") as f:
                                    json.dump(new_cache, f, ensure_ascii=False, indent=2)
                                os.replace(tmp, epsilon_path)
                        log.info("delete_invoices: Deleted %d marks from epsilon cache %s for VAT %s", deleted_from_epsilon, epsilon_path, vat)
                    except Exception:
                        log.exception("delete_invoices: failed to persist epsilon cache after deletion")
                else:
                    log.info("delete_invoices: No matching marks found in epsilon cache %s for deletion.", epsilon_path)
            else:
                log.info("delete_invoices: Epsilon cache does not exist at %s; skipping epsilon deletion.", epsilon_path)
        else:
            log.info("delete_invoices: No active VAT available; skipped epsilon deletion.")
    except Exception:
        log.exception("delete_invoices: Error while deleting from epsilon cache")

    # Final summary
    total_requested = len(marks_to_delete)
    flash(f"Διαγράφηκαν {total_requested} επιλεγμένα mark(s). Αφαιρέθηκαν από Excel: {deleted_from_excel}, από Epsilon cache: {deleted_from_epsilon}", "success")
    log.info("delete_invoices: finished request. requested=%d excel=%d epsilon=%d", total_requested, deleted_from_excel, deleted_from_epsilon)

    return redirect(url_for("search"))










# ---------------- Health ----------------
@app.route("/health")
def health():
    return "OK"

# ---------------- Global error handler ----------------
@app.errorhandler(Exception)
def handle_unexpected_error(e):
    tb = traceback.format_exc()
    log.error("Unhandled exception: %s\n%s", str(e), tb)
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    if debug:
        return "<pre>{}</pre>".format(escape(tb)), 500
    return safe_render("error_generic.html", message="Συνέβη σφάλμα στον server. Δες logs."), 500

@app.route('/favicon.ico')
def favicon():
    return '', 204

if __name__ == "__main__":
    port = int(os.getenv("PORT", "5001"))
    debug_flag = True
    app.run(host="0.0.0.0", port=port, debug=debug_flag, use_reloader=True)